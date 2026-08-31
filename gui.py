# -*- coding: utf-8 -*-
"""拼豆助手 - 桌面版（现代拟态卡片 UI）。
运行入口：python gui.py（打包后运行 exe 即可）"""

import os
import shutil
import sys
import threading
import webbrowser
import tkinter as tk
import tkinter.font as tkfont
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageDraw, ImageTk

import palettes
import bead_engine as be
import inspirations as ins
import member as member_mod
import i18n as L


APP_NAME = "拼豆助手"

# --------------------------------------------------------------------------
# 视觉设计词条 —— 现代拟态（Neumorphism）：
#   柔和中性底 + 凹陷/凸起的同色相阴影，营造“柔软卡片”的质感。
# --------------------------------------------------------------------------
BG        = "#EDF0F4"   # 全局背景（浅灰蓝，拟态基底）
BG_DEEP   = "#E2E6EC"   # 更深一点的基底（用于凹陷区）
CARD      = "#F5F7FA"   # 卡片面板色
PAPER     = "#FFFFFF"   # 画布纸面
INK       = "#2B3440"   # 主文字（深墨蓝）
INK_SOFT  = "#6B7686"   # 次要文字
INK_FAINT = "#9AA4B2"   # 弱提示文字
BEAD      = "#E8483A"   # 豆红：主行动色
TEAL      = "#2E8C83"   # 豆青：次级强调（状态/编号/进度）
LIGHT     = "#FFFFFF"   # 高光
SHADOW    = "#C4CBD6"   # 阴影
FONT      = "Microsoft YaHei UI"


def _hex_lighter(hx, ratio=0.35):
    """颜色与白色混合，返回更浅的 hex（用于渐变/图例底色）。"""
    r = int(hx[1:3], 16); g = int(hx[3:5], 16); b = int(hx[5:7], 16)
    return "#%02X%02X%02X" % (int(r + (255 - r) * ratio),
                              int(g + (255 - g) * ratio),
                              int(b + (255 - b) * ratio))


def resource_path(rel):
    """兼容 PyInstaller 打包后的资源路径。"""
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, rel)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), rel)


def storage_dir():
    """用户数据目录：打包版(exe)与 exe 同级，源码版与脚本同级。"""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def _guide_path():
    """AI 卡通 Key 教程页路径：打包版取 _MEIPASS，源码版取脚本同级。"""
    if getattr(sys, "frozen", False):
        base = getattr(sys, "_MEIPASS", os.path.dirname(sys.executable))
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, "ai_cartoon_guide.html")


import json as _json

def _load_settings():
    try:
        with open(os.path.join(storage_dir(), "user_settings.json"),
                  encoding="utf-8") as f:
            return _json.load(f)
    except Exception:
        return {}


def _save_settings(data):
    try:
        with open(os.path.join(storage_dir(), "user_settings.json"), "w",
                  encoding="utf-8") as f:
            _json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


class RoundedButton(tk.Canvas):
    """圆角按钮（Canvas 绘制，支持悬停反馈与禁用态）。"""

    def __init__(self, master, text, command=None, color=BEAD, fg="#FFFFFF",
                 font=(FONT, 10, "bold"), radius=16, padx=18, pady=9, **kw):
        self._text = text
        self._command = command
        self._color = color
        self._fg = fg
        self._font = font
        self._r = radius
        self._px, self._py = padx, pady
        self._disabled = False
        super().__init__(master, bg=master["bg"], highlightthickness=0,
                         bd=0, cursor="hand2", **kw)
        ft = tkfont.Font(font=font)
        text_w = ft.measure(text)
        self._tw, self._th = text_w, ft.metrics("linespace")
        self.configure(width=text_w + self._px * 2 + 4,
                       height=self._th + self._py * 2 + 4)
        self.bind("<Button-1>", self._clicked)
        self.bind("<Enter>", lambda e: self._draw(hover=True))
        self.bind("<Leave>", lambda e: self._draw(hover=False))
        self._draw()

    def _clicked(self, ev):
        if not self._disabled and self._command:
            self._command()

    def set_text(self, text):
        self._text = text
        ft = tkfont.Font(font=self._font)
        self._tw = ft.measure(text)
        self.configure(width=self._tw + self._px * 2 + 4)
        self._draw()

    def set_state(self, disabled):
        self._disabled = disabled
        self.configure(cursor="arrow" if disabled else "hand2")
        self._draw()

    def _draw(self, hover=False):
        self.delete("all")
        w, h = self.winfo_reqwidth(), self.winfo_reqheight()
        col = self._color
        if hover and not self._disabled:
            col = _hex_lighter(self._color, 0.12)
        if self._disabled:
            col = "#B9C0CB"
            fg = "#FFFFFF"
        else:
            fg = self._fg
        # 柔和投影
        self.create_polygon(self._shadow_pts(w, h), fill="#C9D0DA", outline="")
        self.create_text(3, 4, text="", fill="")
        # 主体圆角矩形
        self._round_rect(3, 3, w - 4, h - 4, self._r, fill=col, outline=col)
        # 顶部高光条，增强拟态立体感
        self._round_rect(3, 3, w - 4, 3 + int(self._th * 0.5), self._r,
                         fill=_hex_lighter(col, 0.10), outline="")
        self.create_text(w / 2 + 1, h / 2 + 1, text=self._text, fill=fg,
                         font=self._font, anchor="center")

    def _round_rect(self, x1, y1, x2, y2, r, **kw):
        pts = [x1 + r, y1, x2 - r, y1, x2, y1, x2, y1 + r, x2, y2 - r, x2, y2,
               x2 - r, y2, x1 + r, y2, x1, y2, x1, y2 - r, x1, y1 + r, x1, y1]
        return self.create_polygon(pts, smooth=True, **kw)

    def _shadow_pts(self, w, h):
        off = 3
        return [off, h - off, w - off, h - off, w - off, off,
                off, off]


class Chip(tk.Canvas):
    """胶囊小按钮（头部次要操作：打开/保存等）。"""

    def __init__(self, master, text, command=None, color="#FFFFFF", fg=INK,
                 font=(FONT, 9, "bold"), **kw):
        self._text = text
        self._command = command
        self._color = color
        self._fg = fg
        self._font = font
        self._hover = False
        super().__init__(master, bg=master["bg"], highlightthickness=0,
                         bd=0, cursor="hand2", **kw)
        ft = tkfont.Font(font=font)
        tw = ft.measure(text)
        th = ft.metrics("linespace")
        self.configure(width=tw + 26, height=th + 12)
        self.bind("<Button-1>", lambda e: self._command() if self._command else None)
        self.bind("<Enter>", lambda e: self._set_hover(True))
        self.bind("<Leave>", lambda e: self._set_hover(False))
        self._draw()

    def set_text(self, text):
        self._text = text
        sf = tkfont.Font(font=self._font)
        sw = sf.measure(text)
        sh = sf.metrics("linespace")
        self.configure(width=sw + 26, height=sh + 12)
        self._draw()

    def _set_hover(self, hv):
        self._hover = hv
        self._draw()

    def _draw(self):
        self.delete("all")
        w, h = self.winfo_reqwidth(), self.winfo_reqheight()
        x1, y1, x2, y2 = 2, 2, w - 2, h - 2
        base = _hex_lighter(self._color, 0.06) if self._hover else self._color
        self.create_polygon([x1 + 14, y1, x2 - 14, y1, x2, y1 + 10, x2, y2 - 10,
                             x2 - 14, y2, x1 + 14, y2, x1, y2 - 10, x1, y1 + 10],
                            fill=base, outline="#D5DBE3", width=1)
        self.create_text(w / 2, h / 2 + 1, text=self._text, fill=self._fg,
                         font=self._font, anchor="center")


class SegmentedProgress(tk.Frame):
    """单条进度条：一条连续圆角轨 + 整体百分比 + 左侧阶段名。

    三个阶段（网格化 / 颜色映射 / 渲染预览）按等权平均算出总进度，
    阶段名只在左上方提示"当前到哪一步"，不再把轨道切成三块，
    避免末段用另一种颜色被误读成警告。

    - set_value(stage_0to2, frac): 推进某一段（0~1），其余段自动补齐
    - set_done(stage)             : 某段完成（100%）
    - reset()                     : 全部归零
    """

    STAGES = (L.tr("网格化"), L.tr("颜色映射"), L.tr("渲染预览"))
    TRACK = "#DCE2EA"
    # 轨道两端留出的圆角半径（轨道本身贴满控件宽度）
    _PAD = 2

    def __init__(self, master, bg=BG, width=320, height=34):
        self._width, self._height = width, height
        self._bg = bg
        super().__init__(master, bg=bg)
        # 上部一行放"阶段名 + 百分比"，下部一条连续轨道
        top = tk.Frame(self, bg=bg)
        top.pack(fill=tk.X, pady=(0, 3))
        self._stage_lbl = tk.Label(top, text="", bg=bg, fg=INK_SOFT,
                                   font=(FONT, 8), anchor=tk.W)
        self._stage_lbl.pack(side=tk.LEFT)
        self._pct_lbl = tk.Label(top, text="", bg=bg, fg=INK_SOFT,
                                 font=(FONT, 8, "bold"), anchor=tk.E)
        self._pct_lbl.pack(side=tk.RIGHT)

        self._canvas = tk.Canvas(self, width=width, height=height, bg=bg,
                                 highlightthickness=0)
        self._canvas.pack()
        y0, y1 = self._PAD, height - self._PAD
        r = max(1, (y1 - y0) // 2)
        x0, x1 = 0, width
        self._track_box = (x0, y0, x1, y1)
        self._canvas.create_polygon(
            self._round_poly(x0 + 1, y0 + 1, x1, y1, r),
            fill=self.TRACK, outline="")
        self._fill_id = self._canvas.create_polygon(
            self._round_poly(x0, y0, x0 + 0.001, y1, r),
            fill=TEAL, outline="")
        self._fracs = [0.0, 0.0, 0.0]

    def _round_poly(self, x1, y1, x2, y2, r):
        return [x1 + r, y1, x2 - r, y1, x2, y1 + 2, x2, y2 - 2, x2 - r, y2,
                x1 + r, y2, x1, y2 - 2, x1, y1 + 2, x1, y1]

    def set_value(self, idx, frac, label=None):
        """推进第 idx 段。只把"之前已经走完的段"补满，不动后续段，
        否则 set_done(0) 会把 3 段全部补满、百分比直接跳到 100%。

        label: 当前阶段名（可选）。总百分比是三段等权平均，但"现在在
        做哪一步"只有调用方知道，所以显式传入，不从 frac 反推。
        """
        idx = max(0, min(len(self.STAGES) - 1, idx))
        frac = max(0.0, min(1.0, float(frac)))
        # 已完成的段（进度到 1）的最大下标；只补它之前的段
        done_max = -1
        for i in range(len(self._fracs)):
            if self._fracs[i] >= 1.0:
                done_max = i
        for i in range(min(idx, done_max)):
            self._fracs[i] = 1.0
        self._fracs[idx] = max(self._fracs[idx], frac)
        if label is not None:
            self._stage_lbl.config(text=label)
        self._paint()

    def set_done(self, idx):
        """第 idx 段完成。若它已跑满则原样返回（阶段名不变，保持上一步
        传入的名字）；若之前没标过进度，显示本段名，避免文字提前切换。"""
        if self._fracs[idx] >= 1.0:
            return
        self.set_value(idx, 1.0, label=self.STAGES[idx])

    def reset(self):
        self._fracs = [0.0] * len(self.STAGES)
        self._stage_lbl.config(text="")
        self._paint()

    def _paint(self):
        pct = sum(self._fracs) / len(self.STAGES)
        x0, y0, x1, y1 = self._track_box
        r = max(1, (y1 - y0) // 2)
        fx = x0 + (x1 - x0) * pct
        self._canvas.coords(
            self._fill_id, *self._round_poly(x0, y0, max(x0 + 0.001, fx), y1, r))
        self._pct_lbl.config(text="%d%%" % round(pct * 100))


def _mk_button(parent, text, command, color=BEAD, fg="#FFFFFF",
               font=(FONT, 10, "bold")):
    return RoundedButton(parent, text, command, color=color, fg=fg, font=font)


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(L.tr("拼豆助手 —— 照片变拼豆图纸"))
        self.geometry("1260x830")
        self.minsize(1080, 720)
        self.configure(bg=BG)

        self.src_img = None
        self.base = None
        self.crop_box = None
        self.result = None
        self.used_colors = []
        self.grid_cols = self.grid_rows = 0
        self.canvas_photo = None
        self.full_img = None
        self.zoom_level = 1.0
        self.canvas_offset_x = 0
        self.canvas_offset_y = 0
        self._dragging = False
        self._drag_start = (0, 0)
        self.overlay_id = None
        self.remove_alpha = False
        self.showing_number = False
        self.highlight_color = None   # 单色分布视图：当前高亮的颜色名(None=全部)
        self._swatch_chips = []       # 色板chips：(name, canvas, outline_item)
        self.done_cells = set()       # 拼制进度：已拼格子集合 (r, c)
        self._done_items = []         # 画布上已拼 X 记号的 canvas item
        # 右画布（拼豆图纸）状态
        self.canvas_right_photo = None
        self.full_img_right = None
        self.zoom_level_right = 1.0
        self.canvas_offset_x_r = 0
        self.canvas_offset_y_r = 0
        self._dragging_right = False
        self._drag_start_right = (0, 0)

        self._last_path = None          # 最近打开图片（供“记住上次设置”）
        self._ai_key_var = tk.StringVar(value="")   # SenseNova API key（本地保存，不入源码）
        self._ai_cartoon_running = False

        # 语言：启动时先读盘（不触发回调），读不到就默认中文
        L.load_lang_from_settings(_load_settings())
        self._lang = L.get_lang()
        # 语言切换后整界面重建（只保留算法数据：图片/图纸/进度，UI 全部重建）
        L.on_lang_changed(self._rebuild_ui_on_lang)

        self._setup_style()
        self._build_ui()
        self.draw_empty_state()
        self.status(L.tr("打开一张照片，把它变成拼豆图纸"))
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self.after(50, self._restore_settings)

    def _rebuild_ui_on_lang(self):
        """语言切换后重建整个界面。数据保留，只重建控件与文案。

        不能只改文字：布局是构建时按文案长度排好的，且 ttk.Radiobutton /
        LabelFrame 的 title 走的是系统主题，textvariable 绑定不生效，
        整界面重建最稳。
        """
        self._lang = L.get_lang()
        try:
            for w in self.winfo_children():
                w.destroy()
        except Exception:
            pass
        # 清空构建期缓存的控件引用，避免指向已销毁的 widget
        self.toggle_btn = None
        self._swatch_chips = []
        self._done_items = []
        self._insp_photos = getattr(self, "_insp_photos", [])
        self._setup_style()
        self._build_ui()
        if self.result is None:
            self.draw_empty_state()
        self.status(L.tr("打开一张照片，把它变成拼豆图纸"))

    def _toggle_lang(self):
        """切换中/英。先真正切换语言（set_lang 触发整界面重建），再写盘记住。"""
        target = "en" if self._lang == "zh" else "zh"
        L.set_lang(target)          # 改模块语言并触发 _rebuild_ui_on_lang 重建界面
        self._lang = L.get_lang()   # 回调已把 self._lang 更新，这里再兜底读回
        try:
            s = _load_settings()
            L.save_lang_to_settings(s, self._lang)
            _save_settings(s)
        except Exception:
            pass
        self.title(L.tr("拼豆助手 —— 照片变拼豆图纸"))

    def _setup_style(self):
        """让 ttk 控件融入拟态主题（底色、字体、字号）。"""
        try:
            self.tk.call("ttk::style", "theme", "vista")
        except Exception:
            pass
        st = ttk.Style(self)
        st.configure(".", font=(FONT, 10))
        st.configure("TRadiobutton", background=CARD, font=(FONT, 9),
                     foreground=INK)
        st.configure("TCheckbutton", background=CARD, font=(FONT, 9),
                     foreground=INK)
        st.configure("TSpinbox", fieldbackground=PAPER, foreground=INK,
                     arrowcolor=INK_SOFT, font=(FONT, 10))
        st.configure("TCombobox", fieldbackground=PAPER, foreground=INK,
                     arrowcolor=INK_SOFT, font=(FONT, 10))
        st.map("TRadiobutton", background=[("active", CARD)])
        st.map("TCheckbutton", background=[("active", CARD)])
        st.map("TSpinbox", fieldbackground=[("disabled", "#EDF0F4")])
        st.map("TCombobox", fieldbackground=[("disabled", "#EDF0F4")])

    # ------------------------------------------------------------------ UI
    def _card(self, parent, title=None, **kw):
        """拟态卡片：浅底 + 柔和内阴影 + 上缘微高光。"""
        card = tk.Frame(parent, bg=CARD, highlightthickness=0, bd=0)
        tk.Frame(card, bg=CARD, height=1, bd=0).pack(side=tk.BOTTOM,
                                                     fill=tk.X, pady=(0, 0))
        # 顶部标题行
        if title:
            head = tk.Frame(card, bg=CARD)
            head.pack(fill=tk.X, pady=(0, 6))
            dot = tk.Frame(head, bg=TEAL, width=4, height=16)
            dot.pack(side=tk.LEFT, padx=(2, 8))
            tk.Label(head, text=title, bg=CARD, fg=INK,
                     font=(FONT, 11, "bold")).pack(side=tk.LEFT)
        return card

    def _build_ui(self):
        # ---- 顶部：深色主头 + 品牌 + 全局操作 ----
        head = tk.Frame(self, bg="#2B3440", height=92)
        head.pack(fill=tk.X)
        head.pack_propagate(False)
        logo = tk.Canvas(head, width=70, height=58, bg="#2B3440",
                         highlightthickness=0)
        logo.pack(side=tk.LEFT, padx=(18, 8), pady=8)
        self._draw_logo(logo)
        ttl = tk.Frame(head, bg="#2B3440")
        ttl.pack(side=tk.LEFT, pady=8)
        tk.Label(ttl, text=L.tr("拼豆助手"), bg="#2B3440", fg="#FFFFFF",
                 font=(FONT, 20, "bold")).pack(anchor=tk.W)
        tk.Label(ttl, text=L.tr("照片 → 像素格子 → 照着拼的图纸"),
                 bg="#2B3440", fg="#A8B1BF", font=(FONT, 9)).pack(anchor=tk.W)
        head_tools = tk.Frame(head, bg="#2B3440")
        head_tools.pack(side=tk.RIGHT, padx=18, pady=16, fill=tk.X)
        tools = tk.Frame(head_tools, bg="#2B3440")
        tools.pack(side=tk.LEFT, fill=tk.X, expand=True)
        Chip(tools, L.tr("打开图片…"), self.open_image,
             color="#FFFFFF", fg="#2B3440").pack(side=tk.LEFT)
        Chip(tools, L.tr("保存图纸 PNG"), self.save_pattern,
             color="#FFFFFF", fg="#2B3440").pack(side=tk.LEFT, padx=8)
        Chip(tools, L.tr("保存耗材清单"), self.save_list,
             color="#FFFFFF", fg="#2B3440").pack(side=tk.LEFT, padx=(0, 8))
        Chip(tools, L.tr("导出 PDF"), self.export_pdf,
             color="#FFFFFF", fg="#2B3440").pack(side=tk.LEFT, padx=(0, 8))
        Chip(tools, L.tr("拼豆灵感"), self.open_inspiration,
             color=BEAD, fg="#FFFFFF", font=(FONT, 9, "bold")).pack(
                 side=tk.LEFT, padx=(0, 8))
        self.toggle_btn = Chip(tools, L.tr("编号图 ▶"), self.toggle_view,
                               color=BEAD, fg="#FFFFFF")
        self.toggle_btn.pack(side=tk.LEFT)
        # 语言切换（与网页版一致：🌐 图标 + 中文/EN）
        Chip(tools, "🌐 " + (L.tr("中文") if self._lang == "zh" else "EN"),
             self._toggle_lang, color="#3D4858", fg="#FFFFFF",
             font=(FONT, 8, "bold")).pack(side=tk.LEFT, padx=(10, 0))
        # 缩放工具条（挪到顶栏，不再夹在画布中间）
        zbar = tk.Frame(head_tools, bg="#2B3440")
        zbar.pack(side=tk.LEFT, padx=(12, 0))
        self.zoom_out_btn = Chip(zbar, "－", lambda: self.zoom(0.85),
                                 color="#3D4858", fg="#FFFFFF")
        self.zoom_out_btn.pack(side=tk.LEFT, padx=(0, 4))
        self.zoom_label = tk.Label(zbar, text="100%", bg="#2B3440", fg="#FFFFFF",
                                   font=(FONT, 9, "bold"), width=6)
        self.zoom_label.pack(side=tk.LEFT)
        self.zoom_in_btn = Chip(zbar, "＋", lambda: self.zoom(1.18),
                                color="#3D4858", fg="#FFFFFF")
        self.zoom_in_btn.pack(side=tk.LEFT, padx=(4, 6))
        Chip(zbar, L.tr("适应画布"), self.zoom_fit, color=TEAL,
             fg="#FFFFFF").pack(side=tk.LEFT)
        Chip(zbar, L.tr("⟷ 翻转"), lambda: self.flip_view("h"),
             color="#3D4858", fg="#A8B1BF", font=(FONT, 8, "bold")).pack(
                 side=tk.LEFT, padx=(6, 0))
        Chip(zbar, L.tr("↺ 重置"), self.reset_progress,
             color="#3D4858", fg="#A8B1BF", font=(FONT, 8, "bold")).pack(
                 side=tk.LEFT, padx=(6, 0))
        Chip(zbar, L.tr("⟳ 垂翻"), lambda: self.flip_view("v"),
             color="#3D4858", fg="#A8B1BF", font=(FONT, 8, "bold")).pack(
                 side=tk.LEFT, padx=(6, 0))

        # ---- 主体 ----
        body = tk.Frame(self, bg=BG, padx=16, pady=14)
        body.pack(fill=tk.BOTH, expand=True)

        # 左：双画布（原图 | 拼豆图纸）
        left = tk.Frame(body, bg=BG)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 左画布容器（原图/抠图）
        left_card = tk.Frame(left, bg=PAPER)
        left_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 3))
        tk.Label(left_card, text=L.tr("原图"), bg=PAPER, fg=INK_FAINT,
                 font=(FONT, 9), anchor=tk.W).pack(fill=tk.X, padx=6, pady=(2, 0))
        self.canvas = tk.Canvas(left_card, bg=PAPER, highlightthickness=0, width=400, height=600)
        self.canvas.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        self.canvas.bind("<Button-1>", self.on_canvas_click)
        self.canvas.bind("<B1-Motion>", self.on_canvas_drag)
        self.canvas.bind("<Button-3>", lambda e: self.clear_crop())
        self.canvas.bind("<MouseWheel>", self.on_mousewheel)
        self.canvas.bind("<Control-MouseWheel>", self.on_mousewheel)
        self.canvas.bind("<Button-2>", self.on_canvas_press)
        self.canvas.bind("<B2-Motion>", self.on_canvas_drag_move)
        self.canvas.bind("<ButtonRelease-2>", self.on_canvas_release)
        self.canvas.bind("<Control-Button-1>", self.on_canvas_press)
        self.canvas.bind("<Control-B1-Motion>", self.on_canvas_drag_move)
        self.canvas.bind("<Control-ButtonRelease-1>", self.on_canvas_release)

        # 右画布容器（拼豆图纸）
        right_card = tk.Frame(left, bg=PAPER)
        right_card.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(3, 0))
        tk.Label(right_card, text="拼豆图纸", bg=PAPER, fg=INK_FAINT,
                 font=(FONT, 9), anchor=tk.W).pack(fill=tk.X, padx=6, pady=(2, 0))
        self.canvas_right = tk.Canvas(right_card, bg=PAPER, highlightthickness=0,
                                      width=400, height=600)
        self.canvas_right.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        self.canvas_right.bind("<MouseWheel>", self.on_mousewheel_right)
        self.canvas_right.bind("<Control-MouseWheel>", self.on_mousewheel_right)
        self.canvas_right.bind("<Button-2>", self.on_canvas_press_right)
        self.canvas_right.bind("<B2-Motion>", self.on_canvas_drag_move_right)
        self.canvas_right.bind("<ButtonRelease-2>", self.on_canvas_release_right)
        self.canvas_right.bind("<Control-Button-1>", self.on_canvas_press_right)
        self.canvas_right.bind("<Control-B1-Motion>", self.on_canvas_drag_move_right)
        self.canvas_right.bind("<Control-ButtonRelease-1>", self.on_canvas_release_right)

        # 缩放工具条已移到顶栏

        # 单色分布条：移到右侧面板，不在画布中间

        # 右：控制台卡片（带滚动条，卡片多了也能上下滚）
        right_outer = tk.Frame(body, bg=BG, width=340)
        right_outer.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(14, 0))
        right_outer.pack_propagate(False)
        self.right_canvas = tk.Canvas(right_outer, bg=BG, highlightthickness=0)
        right_canvas = self.right_canvas
        right_scroll = ttk.Scrollbar(right_outer, orient="vertical",
                                     command=right_canvas.yview)
        right_canvas.configure(yscrollcommand=right_scroll.set)
        right_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        right_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        inner = tk.Frame(right_canvas, bg=BG)
        right_win = right_canvas.create_window((0, 0), window=inner, anchor=tk.NW)
        # 让 inner 跟随 canvas 宽度，避免右侧留空。
        # ★ 修复：不能只 inner.configure(width=...)，pack 传播下这个设置无效，
        #   卡片列会一直停留在 306 宽、面板右侧空出一大片。
        #   直接设置 canvas 窗口项(item)的宽度才能强制 inner 拉伸到视口同宽。
        def _sync_inner_scrollregion(e):
            right_canvas.itemconfigure(right_win, width=e.width)
            inner.configure(width=e.width)
            right_canvas.update_idletasks()
            right_canvas.configure(scrollregion=right_canvas.bbox("all"))
        right_canvas.bind("<Configure>", _sync_inner_scrollregion)

        # ① 抠图
        g1 = self._card(inner, L.tr("① 抠图"))
        g1.pack(fill=tk.X, pady=(2, 8))
        self.mode = tk.StringVar(value="none")
        mode_desc = {"auto": L.tr("自动去背景（推荐 · 纯色背景）"),
                     "manual": L.tr("手动框选主体（点两点，右键取消）"),
                     "none": L.tr("不抠图（整张都用）")}
        for key, txt in mode_desc.items():
            ttk.Radiobutton(g1, text=txt, value=key, variable=self.mode,
                            command=self.apply_cutout).pack(anchor=tk.W, pady=1)

        # ② 底板
        g2 = self._card(inner, L.tr("② 底板"))
        g2.pack(fill=tk.X, pady=(2, 8))
        row = tk.Frame(g2, bg=CARD)
        row.pack(fill=tk.X, pady=2)
        tk.Label(row, text=L.tr("宽"), bg=CARD, fg=INK, font=(FONT, 9)).pack(side=tk.LEFT)
        self.lW = ttk.Spinbox(row, from_=8, to=300, width=5)
        self.lW.set(29); self.lW.pack(side=tk.LEFT, padx=4)
        tk.Label(row, text=L.tr("高"), bg=CARD, fg=INK, font=(FONT, 9)).pack(side=tk.LEFT)
        self.lH = ttk.Spinbox(row, from_=8, to=300, width=5)
        self.lH.set(29); self.lH.pack(side=tk.LEFT, padx=4)
        chip_row = tk.Frame(g2, bg=CARD)
        chip_row.pack(fill=tk.X, pady=(4, 2))
        for label, fn in (("29×29", lambda: self.set_grid(29, 29)),
                          ("45×45", lambda: self.set_grid(45, 45)),
                          (L.tr("按图片推荐"), self.suggest_dialog)):
            Chip(chip_row, label, fn, color="#E9EDF3", fg=INK_SOFT,
                 font=(FONT, 8, "bold")).pack(side=tk.LEFT, padx=(0, 4))
        self.keep_ratio = tk.BooleanVar(value=False)
        ttk.Checkbutton(g2, text=L.tr("按原图比例自动调整高"),
                        variable=self.keep_ratio).pack(anchor=tk.W, pady=(4, 0))
        tk.Label(g2, text=L.tr("29×29 / 45×45=标准板；宽度决定高度（等比）"),
                 bg=CARD, fg=INK_FAINT, font=(FONT, 8)).pack(anchor=tk.W, pady=(2, 0))

        # ③ 色卡
        g3 = self._card(inner, L.tr("③ 色卡"))
        g3.pack(fill=tk.X, pady=(2, 8))
        row = tk.Frame(g3, bg=CARD); row.pack(fill=tk.X)
        tk.Label(row, text=L.tr("色卡"), bg=CARD, fg=INK, font=(FONT, 9)).pack(side=tk.LEFT)
        self.pal_var = tk.StringVar(value=palettes.get_all_palettes()[0]["name"])
        self.pal_box = ttk.Combobox(row, textvariable=self.pal_var, state="readonly",
                                    values=[p["name"] for p in palettes.get_all_palettes()],
                                    width=15)
        self.pal_box.pack(side=tk.LEFT, padx=4)
        Chip(row, L.tr("色卡编辑"), self.edit_palette, color="#E9EDF3", fg=INK_SOFT,
             font=(FONT, 8, "bold")).pack(side=tk.LEFT, padx=(4, 0))
        row2 = tk.Frame(g3, bg=CARD); row2.pack(fill=tk.X, pady=(6, 0))
        tk.Label(row2, text=L.tr("最多颜色数"), bg=CARD, fg=INK, font=(FONT, 9)).pack(side=tk.LEFT)
        self.maxC = ttk.Spinbox(row2, from_=2, to=80, width=4)
        self.maxC.set(12); self.maxC.pack(side=tk.LEFT, padx=4)
        tk.Label(g3, text=L.tr("数字越小越省豆省事，越大越还原照片"),
                 bg=CARD, fg=INK_FAINT, font=(FONT, 8)).pack(anchor=tk.W, pady=(4, 0))

        # ④ 空白格
        g4 = self._card(inner, L.tr("④ 空白格"))
        g4.pack(fill=tk.X, pady=(2, 8))
        self.bg_fill = tk.BooleanVar(value=False)
        ttk.Checkbutton(g4, text=L.tr("空白格用白色豆填满（铺满底板）"),
                        variable=self.bg_fill).pack(anchor=tk.W)

        # ⑤ 卡通化
        g5 = self._card(inner, L.tr("⑤ 卡通化"))
        g5.pack(fill=tk.X, pady=(2, 8))
        self.cartoon_enabled = tk.BooleanVar(value=False)
        ttk.Checkbutton(g5, text=L.tr("开启卡通化（粗黑边 + 平涂色块，更适合拼豆）"),
                        variable=self.cartoon_enabled).pack(anchor=tk.W)
        row5 = tk.Frame(g5, bg=CARD)
        row5.pack(fill=tk.X, pady=(4, 2))
        tk.Label(row5, text=L.tr("边缘"), bg=CARD, fg=INK, font=(FONT, 9)).pack(side=tk.LEFT)
        self.cartoon_edge = ttk.Spinbox(row5, from_=1, to=5, width=3)
        self.cartoon_edge.set(3); self.cartoon_edge.pack(side=tk.LEFT, padx=4)
        tk.Label(row5, text=L.tr("平滑"), bg=CARD, fg=INK, font=(FONT, 9)).pack(side=tk.LEFT, padx=(10, 0))
        self.cartoon_smooth = ttk.Spinbox(row5, from_=1, to=5, width=3)
        self.cartoon_smooth.set(3); self.cartoon_smooth.pack(side=tk.LEFT, padx=4)
        self.cartoon_status = tk.Label(row5, text="", bg=CARD, fg=TEAL, font=(FONT, 8))
        self.cartoon_status.pack(side=tk.LEFT, padx=8)
        Chip(row5, L.tr("检测主体"), self._detect_subject, color="#E9EDF3",
             fg=INK_SOFT, font=(FONT, 8, "bold")).pack(side=tk.LEFT, padx=(6, 0))
        tk.Label(g5, text=L.tr("边缘越粗线条越突出，平滑越强颜色越平"),
                 bg=CARD, fg=INK_FAINT, font=(FONT, 8)).pack(anchor=tk.W, pady=(2, 0))

        # AI 卡通（SenseNova 图像编辑，真·AI 生成）
        tk.Frame(g5, bg="#C9D0DA", height=1).pack(fill=tk.X, pady=(8, 4))
        tk.Label(g5, text=L.tr("AI 卡通（调用商汤 SenseNova，需联网）"),
                 bg=CARD, fg="#B9842C", font=(FONT, 9, "bold")).pack(anchor=tk.W)
        krow = tk.Frame(g5, bg=CARD); krow.pack(fill=tk.X, pady=(3, 2))
        tk.Label(krow, text="Key", bg=CARD, fg=INK_SOFT, font=(FONT, 8)).pack(side=tk.LEFT)
        self._ai_key_entry = ttk.Entry(krow, textvariable=self._ai_key_var, width=20, show="●")
        self._ai_key_entry.pack(side=tk.LEFT, padx=(4, 4), fill=tk.X, expand=True)
        Chip(krow, L.tr("保存"), self._save_ai_key, color="#E9EDF3",
             fg=INK_SOFT, font=(FONT, 8)).pack(side=tk.LEFT, padx=(2, 0))
        self._ai_guide_link = tk.Label(g5, text=L.tr("🔑 获取 Key 教程 →（点击打开）"),
                                       bg=CARD, fg="#2F6DB0",
                                       font=(FONT, 8, "underline"), cursor="hand2")
        self._ai_guide_link.pack(anchor=tk.W, pady=(0, 2))
        self._ai_guide_link.bind("<Button-1>", self._open_ai_guide)
        promrow = tk.Frame(g5, bg=CARD); promrow.pack(fill=tk.X, pady=(2, 2))
        tk.Label(promrow, text=L.tr("风格"), bg=CARD, fg=INK_SOFT, font=(FONT, 8)).pack(side=tk.LEFT)
        self._ai_prompt_var = tk.StringVar(
            value="把它变成可爱卡通风格，Q版动漫，色彩明亮")
        self._ai_prompt_entry = ttk.Entry(promrow, textvariable=self._ai_prompt_var,
                                          width=24)
        self._ai_prompt_entry.pack(side=tk.LEFT, padx=(4, 0))
        brow = tk.Frame(g5, bg=CARD); brow.pack(fill=tk.X, pady=(4, 0))
        self._ai_btn = RoundedButton(brow, L.tr("AI 卡通化"), self._ai_cartoon,
                                     color="#8E6BDB", fg="#FFFFFF",
                                     font=(FONT, 10, "bold"))
        self._ai_btn.pack(side=tk.LEFT)
        self._ai_status = tk.Label(brow, text="", bg=CARD, fg=TEAL, font=(FONT, 8))
        self._ai_status.pack(side=tk.LEFT, padx=8)
        tk.Label(g5, text=L.tr("生成卡通图后替换原图，再点「生成图纸」即可"),
                 bg=CARD, fg=INK_FAINT, font=(FONT, 8)).pack(anchor=tk.W, pady=(2, 0))

        # ⑥ 会员（按钮，点开弹窗显示收款码 + 兑换码激活）
        g6 = self._card(inner, L.tr("💎 会员"))
        g6.pack(fill=tk.X, pady=(2, 8))
        self._member_status_var = tk.StringVar(value=L.tr("未开通会员"))
        self._member_status_label = tk.Label(g6, textvariable=self._member_status_var,
                 bg=CARD, fg="#8A93A0", font=(FONT, 9, "bold"),
                 anchor=tk.W)
        self._member_status_label.pack(fill=tk.X)
        _mrow = tk.Frame(g6, bg=CARD); _mrow.pack(fill=tk.X, pady=(4, 0))
        self._open_member_btn = Chip(_mrow, L.tr("💳 开通 / 激活会员"),
                                     self._open_member_dialog,
                                     color="#8E6BDB", fg="#FFFFFF",
                                     font=(FONT, 9, "bold"))
        self._open_member_btn.pack(fill=tk.X)
        tk.Label(g6, text=member_mod.PRICE,
                 bg=CARD, fg="#B9842C", font=(FONT, 8, "bold"),
                 anchor=tk.W).pack(anchor=tk.W, pady=(2, 0))

        # 刷新会员状态显示
        self._refresh_member_display()

        # 主 CTA + 进度
        self.gen_btn = RoundedButton(inner, L.tr("生 成 图 纸"), self.generate,
                                     color=BEAD, fg="#FFFFFF",
                                     font=(FONT, 13, "bold"))
        self.gen_btn.pack(fill=tk.X, pady=(4, 10))

        # 单色分布：可折叠小面板（点标题展开/收起；生成后显示标题条）
        self.swatch_card = tk.Frame(inner, bg=CARD, bd=1, relief="solid",
                                    highlightthickness=0,
                                    highlightbackground="#C9D0DA")
        self.swatch_head = tk.Frame(self.swatch_card, bg=CARD, cursor="hand2")
        self.swatch_head.pack(fill=tk.X, padx=0, pady=6)
        self._swatch_open = False
        self.swatch_arrow = tk.Label(self.swatch_head, text="▶",
                                     bg=CARD, fg=INK_SOFT, font=(FONT, 8, "bold"),
                                     anchor=tk.W)
        self.swatch_arrow.pack(side=tk.LEFT, padx=(8, 4))
        self.swatch_head.bind("<Button-1>", self._toggle_swatches)
        self._swatch_title = tk.Label(self.swatch_head, text=L.tr("单色分布"),
                 bg=CARD, fg=INK_SOFT, font=(FONT, 9, "bold"),
                 cursor="hand2")
        self._swatch_title.pack(side=tk.LEFT)
        tk.Label(self.swatch_head, text=L.tr("（点开只看该色格子）"),
                 bg=CARD, fg=INK_FAINT, font=(FONT, 7),
                 cursor="hand2").pack(side=tk.LEFT, padx=(2, 0))
        self.swatch_canvas = tk.Canvas(self.swatch_card, bg=CARD, height=58,
                                       highlightthickness=0)
        self.swatch_inner = tk.Frame(self.swatch_canvas, bg=CARD)
        self.swatch_canvas.create_window((0, 0), window=self.swatch_inner,
                                         anchor=tk.NW)
        self.swatch_inner.bind(
            "<Configure>",
            lambda e: self.swatch_canvas.configure(
                scrollregion=self.swatch_canvas.bbox("all")))
        self.swatch_canvas.bind("<MouseWheel>", self._on_swatch_wheel)
        self.swatch_canvas.bind("<Shift-MouseWheel>", self._on_swatch_wheel)
        self.swatch_card.pack(fill=tk.X, pady=(2, 4))
        # 初始就显示标题条；颜色区默认收起，生成后填充

        self.prog = SegmentedProgress(inner, bg=BG, width=306, height=22)
        self.prog.pack(fill=tk.X, pady=(0, 8))

        self.stats_var = tk.StringVar(value="")
        stats = tk.Label(inner, textvariable=self.stats_var, justify=tk.LEFT,
                         font=(FONT, 9), bg=BG, fg=INK_SOFT, anchor=tk.W)
        stats.pack(anchor=tk.W, fill=tk.X, pady=(0, 4))

        tk.Label(inner, text=L.tr("生成后可在「编号图」与「颜色图」间切换预览"),
                 bg=BG, fg=INK_FAINT, font=(FONT, 8), anchor=tk.W
                 ).pack(anchor=tk.W, fill=tk.X)

        # 滚轮滚右侧面板（★ 修复）：之前只绑在 canvas 上，鼠标停在
        # inner 里的任意卡片/按钮上时事件到不了 canvas，滚轮就没反应。
        # 现在递归给 inner 全部子控件都绑上，唯独跳过单色分布的 swatch_canvas
        # （它自己处理横向滚动）。绑定放在 _build_ui 末尾，确保所有卡片都已创建。
        def _right_panel_wheel(ev):
            self.right_canvas.yview_scroll(-int(ev.delta / 120), "units")
            return "break"

        def _bind_panel_wheel(w):
            if w is getattr(self, "swatch_canvas", None):
                return
            if not w.bind("<MouseWheel>"):
                w.bind("<MouseWheel>", _right_panel_wheel, add="+")
            for c in w.winfo_children():
                _bind_panel_wheel(c)

        _bind_panel_wheel(inner)
        # 内容下方露出的 canvas 空白条也要能滚
        if not right_canvas.bind("<MouseWheel>"):
            right_canvas.bind("<MouseWheel>", _right_panel_wheel, add="+")

        # ---- 底部状态条 ----
        bar = tk.Frame(self, bg="#2B3440", height=30)
        bar.pack(fill=tk.X)
        bar.pack_propagate(False)
        dot = tk.Canvas(bar, width=20, height=28, bg="#2B3440",
                        highlightthickness=0)
        dot.pack(side=tk.LEFT, padx=(14, 4))
        dot.create_oval(6, 8, 14, 16, fill=TEAL, outline="")
        self.status_var = tk.StringVar(value="")
        tk.Label(bar, textvariable=self.status_var, bg="#2B3440", fg="#A8B1BF",
                 font=(FONT, 9), anchor=tk.W).pack(side=tk.LEFT, fill=tk.X,
                                                   expand=True, pady=4)

    # ---------------------------------------------------------- 豆子 Logo
    def _draw_logo(self, cv):
        c = "#3A4554"
        r = "#E8483A"
        r2 = "#C93629"
        heart = [
            "..rr......rr..",
            ".rRRr....rRRr.",
            ".rRRRr..rRRRr.",
            "..rRRRRRRRRr..",
            "...rRRRRRRr...",
            "....rRRRRr....",
            "......rRr.....",
            ".......r......",
        ]
        s = 4.6
        ox, oy = 7, 6
        for yi, line in enumerate(heart):
            for xi, ch_ in enumerate(line):
                col = c if ch_ == "." else r
                x, y = ox + xi * s, oy + yi * s
                cv.create_rectangle(x, y, x + s - 0.6, y + s - 0.6, fill=col,
                                    outline="" if ch_ == "." else r2)

    # ---------------------------------------------------------- 空态引导
    def _empty_canvas(self, canvas, w, h, title, sub):
        canvas.delete("all")
        step = 36
        for x in range(0, w, step):
            canvas.create_line(x, 0, x, h, fill="#F1F3F7")
        for y in range(0, h, step):
            canvas.create_line(0, y, w, y, fill="#F1F3F7")
        canvas.create_text(w / 2, h / 2 - 20, text=title,
                           fill="#AAB3C0", font=(FONT, 18, "bold"))
        canvas.create_text(w / 2, h / 2 + 14, text=sub,
                           fill="#B8C0CC", font=(FONT, 10))

    def draw_empty_state(self):
        self.canvas.delete("all")
        self.canvas_item = None
        self.canvas_right.delete("all")
        w1 = self.canvas.winfo_width() or 400
        h1 = self.canvas.winfo_height() or 600
        w2 = self.canvas_right.winfo_width() or 400
        h2 = self.canvas_right.winfo_height() or 600
        self._empty_canvas(self.canvas, w1, h1, L.tr("还没有图片"),
                           L.tr("点击左上角「打开图片…」选择一张照片\n推荐纯色背景 · 主体清晰"))
        self._empty_canvas(self.canvas_right, w2, h2, L.tr("拼豆图纸"),
                           L.tr("点「生成图纸」后图纸会显示在这里\n滚轮缩放 · 按住滚轮拖动"))

    # ---------------------------------------------------------- 状态
    def set_status(self, text):
        self.status_var.set(text)
        self.update_idletasks()

    status = set_status

    # ---------------------------------------------------------- 画布显示
    def show_pil(self, im):
        self.full_img = im
        self.zoom_level = 1.0
        self.canvas_offset_x = 0
        self.canvas_offset_y = 0
        self._apply_zoom()

    def _apply_zoom(self):
        if self.full_img is None:
            return
        im = self.full_img
        cw, ch = self.canvas.winfo_width(), self.canvas.winfo_height()
        if cw < 40:
            cw, ch = 820, 640
        base_ratio = min((cw - 8) / im.width, (ch - 8) / im.height)
        ratio = base_ratio * self.zoom_level
        new_w = max(1, int(im.width * ratio))
        new_h = max(1, int(im.height * ratio))
        self.canvas_scalex = ratio
        self.canvas_scaley = ratio
        self.canvas_width = new_w
        self.canvas_height = new_h
        disp = im.resize((new_w, new_h))
        bg = Image.new("RGBA", disp.size, (255, 255, 255))
        bg.alpha_composite(disp.convert("RGBA"))
        self.canvas.delete("all")
        self.canvas_item = None
        self.canvas_photo = ImageTk.PhotoImage(bg)
        x = max(4, (cw - new_w) // 2) + self.canvas_offset_x
        y = max(4, (ch - new_h) // 2) + self.canvas_offset_y
        self.canvas_item = self.canvas.create_image(x, y, anchor=tk.NW,
                                                    image=self.canvas_photo)
        # 记录图片锚点与格宽，供“点击格子标已拼”换算
        self._img_x0 = x
        self._img_y0 = y
        self._cell_px = (new_w / self.grid_cols) if self.grid_cols else 0
        self._cell_py = (new_h / self.grid_rows) if self.grid_rows else 0
        # 缩放工具条更新
        self.zoom_label.config(text=f"{int(self.zoom_level * 100)}%")
        self.clear_crop()
        self._redraw_done_marks()

    def zoom(self, factor):
        new_zoom = max(0.1, min(10.0, self.zoom_level * factor))
        self.zoom_level = new_zoom
        self._apply_zoom()

    def zoom_fit(self):
        self.zoom_level = 1.0
        self.canvas_offset_x = 0
        self.canvas_offset_y = 0
        self._apply_zoom()

    # ---------------------------------------------------------- 右画布（拼豆图纸）
    def show_pil_right(self, im):
        """把 PIL 图像显示在右画布。"""
        self.full_img_right = im
        self.zoom_level_right = 1.0
        self.canvas_offset_x_r = 0
        self.canvas_offset_y_r = 0
        self._apply_zoom_right()

    def _apply_zoom_right(self):
        if self.full_img_right is None:
            return
        im = self.full_img_right
        cw = self.canvas_right.winfo_width() or 400
        ch = self.canvas_right.winfo_height() or 600
        base_ratio = min((cw - 8) / im.width, (ch - 8) / im.height)
        ratio = base_ratio * self.zoom_level_right
        new_w = max(1, int(im.width * ratio))
        new_h = max(1, int(im.height * ratio))
        disp = im.resize((new_w, new_h))
        self.canvas_right_photo = ImageTk.PhotoImage(disp)
        self.canvas_right.delete("all")
        x = max(4, (cw - new_w) // 2) + self.canvas_offset_x_r
        y = max(4, (ch - new_h) // 2) + self.canvas_offset_y_r
        self.canvas_right.create_image(x, y, anchor=tk.NW, image=self.canvas_right_photo)
        self.zoom_label.config(text=f"{int(self.zoom_level_right * 100)}%")

    def on_mousewheel_right(self, ev):
        if self.full_img_right is None:
            return
        new_zoom = max(0.1, min(10.0, self.zoom_level_right * (1.1 if ev.delta > 0 else 0.9)))
        im = self.full_img_right
        cw = self.canvas_right.winfo_width() or 400
        ch = self.canvas_right.winfo_height() or 600
        base_ratio = min((cw - 8) / im.width, (ch - 8) / im.height)
        old_ratio = base_ratio * self.zoom_level_right
        new_ratio = base_ratio * new_zoom
        old_w = max(1, int(im.width * old_ratio))
        old_h = max(1, int(im.height * old_ratio))
        new_w = max(1, int(im.width * new_ratio))
        new_h = max(1, int(im.height * new_ratio))
        old_x0 = max(4, (cw - old_w) // 2) + self.canvas_offset_x_r
        old_y0 = max(4, (ch - old_h) // 2) + self.canvas_offset_y_r
        rx = (ev.x - old_x0) / max(1e-6, old_w)
        ry = (ev.y - old_y0) / max(1e-6, old_h)
        new_x0 = max(4, (cw - new_w) // 2)
        new_y0 = max(4, (ch - new_h) // 2)
        self.canvas_offset_x_r = ev.x - new_x0 - int(rx * new_w)
        self.canvas_offset_y_r = ev.y - new_y0 - int(ry * new_h)
        self.zoom_level_right = new_zoom
        self._apply_zoom_right()

    def on_canvas_press_right(self, ev):
        if self.full_img_right is None:
            return
        self._dragging_right = True
        cw = self.canvas_right.winfo_width() or 400
        ch = self.canvas_right.winfo_height() or 600
        im = self.full_img_right
        base_ratio = min((cw - 8) / im.width, (ch - 8) / im.height)
        ratio = base_ratio * self.zoom_level_right
        img_w = max(1, int(im.width * ratio))
        img_h = max(1, int(im.height * ratio))
        x0 = max(4, (cw - img_w) // 2)
        y0 = max(4, (ch - img_h) // 2)
        self._drag_start_right = (ev.x - self.canvas_offset_x_r - x0,
                                  ev.y - self.canvas_offset_y_r - y0)
        self.canvas_right.config(cursor="fleur")

    def on_canvas_drag_move_right(self, ev):
        if not self._dragging_right:
            return
        cw = self.canvas_right.winfo_width() or 400
        ch = self.canvas_right.winfo_height() or 600
        im = self.full_img_right
        base_ratio = min((cw - 8) / im.width, (ch - 8) / im.height)
        ratio = base_ratio * self.zoom_level_right
        img_w = max(1, int(im.width * ratio))
        img_h = max(1, int(im.height * ratio))
        x0 = max(4, (cw - img_w) // 2)
        y0 = max(4, (ch - img_h) // 2)
        self.canvas_offset_x_r = ev.x - x0 - self._drag_start_right[0]
        self.canvas_offset_y_r = ev.y - y0 - self._drag_start_right[1]
        self._apply_zoom_right()

    def on_canvas_release_right(self, ev):
        if self._dragging_right:
            self._dragging_right = False
            self.canvas_right.config(cursor="arrow")

    def on_mousewheel(self, ev):
        if self.full_img is None:
            return
        self._zoom_anchor(1.1 if ev.delta > 0 else 0.9, ev.x, ev.y)

    def _zoom_anchor(self, factor, mx, my):
        new_zoom = max(0.1, min(10.0, self.zoom_level * factor))
        im = self.full_img
        cw, ch = self.canvas.winfo_width(), self.canvas.winfo_height()
        if cw < 40:
            cw, ch = 820, 640
        base_ratio = min((cw - 8) / im.width, (ch - 8) / im.height)
        old_ratio = base_ratio * self.zoom_level
        new_ratio = base_ratio * new_zoom
        old_w = max(1, int(im.width * old_ratio))
        old_h = max(1, int(im.height * old_ratio))
        new_w = max(1, int(im.width * new_ratio))
        new_h = max(1, int(im.height * new_ratio))
        old_x0 = max(4, (cw - old_w) // 2) + self.canvas_offset_x
        old_y0 = max(4, (ch - old_h) // 2) + self.canvas_offset_y
        rx = (mx - old_x0) / max(1e-6, old_w)
        ry = (my - old_y0) / max(1e-6, old_h)
        new_x0 = max(4, (cw - new_w) // 2)
        new_y0 = max(4, (ch - new_h) // 2)
        self.canvas_offset_x = mx - new_x0 - int(rx * new_w)
        self.canvas_offset_y = my - new_y0 - int(ry * new_h)
        self.zoom_level = new_zoom
        self._apply_zoom()

    # ---------------------------------------------------------- 拖动
    def on_canvas_press(self, ev):
        if self.full_img is None:
            return
        self._dragging = True
        x0 = max(4, (self.canvas.winfo_width() - self.canvas_width) // 2)
        y0 = max(4, (self.canvas.winfo_height() - self.canvas_height) // 2)
        self._drag_start = (ev.x - self.canvas_offset_x - x0,
                            ev.y - self.canvas_offset_y - y0)
        self.canvas.config(cursor="fleur")

    def on_canvas_drag_move(self, ev):
        if not self._dragging:
            return
        x0 = max(4, (self.canvas.winfo_width() - self.canvas_width) // 2)
        y0 = max(4, (self.canvas.winfo_height() - self.canvas_height) // 2)
        self.canvas_offset_x = ev.x - x0 - self._drag_start[0]
        self.canvas_offset_y = ev.y - y0 - self._drag_start[1]
        self._apply_zoom()

    def on_canvas_release(self, ev):
        if self._dragging:
            self._dragging = False
            self.canvas.config(cursor="arrow")

    # ---------------------------------------------------------- 图片与抠图
    def open_image(self, path=None, suggest=True):
        """path=None 时弹文件选择框；suggest=False 用于启动时静默恢复最近图片。"""
        if path is None:
            path = filedialog.askopenfilename(
                title=L.tr("选择照片"),
                filetypes=[(L.tr("图片文件"), "*.png *.jpg *.jpeg *.bmp *.webp"),
                           (L.tr("所有文件"), "*.*")])
            if not path:
                return
        try:
            self.src_img = be.load_image(path)
        except Exception as e:
            if path:
                messagebox.showerror(APP_NAME, f"{L.tr('图片打开失败：')}{e}")
            return
        self._last_path = path
        self.crop_box = None
        self.result = None
        self.apply_cutout(reinteractive=True)
        if suggest:
            self.suggest_dialog(auto=True)

    def apply_cutout(self, reinteractive=False):
        if self.src_img is None:
            return
        m = self.mode.get()
        try:
            if m == "auto":
                self.base = be.remove_uniform_background(self.src_img)
            else:
                self.base = self.src_img.copy()
        except Exception as e:
            messagebox.showwarning(APP_NAME, L.tr("自动去背景失败（退回原图）：") + str(e))
            self.base = self.src_img.copy()
        self.remove_alpha = (m == "auto")
        self.highlight_color = None
        try:
            self.swatch_canvas.pack_forget()
            self._swatch_open = False
            self.swatch_arrow.config(text="▶")
        except Exception:
            pass
        self.show_pil(self.base)
        self.status(L.tr("抠图完成；调整参数后点「生成图纸」。手动模式下：点击画布两次框选主体"))

    # ---------------------------------------------------------- 手动裁剪
    def _to_img(self, ev):
        x = (ev.x - 4) / max(1e-6, self.canvas_scalex)
        y = (ev.y - 4) / max(1e-6, self.canvas_scaley)
        return (max(0, min(int(x), self.base.width - 1)),
                max(0, min(int(y), self.base.height - 1)))

    def on_canvas_click(self, ev):
        if self.mode.get() != "manual":
            # 非手动模式下，生成图纸后：点击格子标记“已拼”
            if self.result is not None and self.grid_cols and self.grid_rows:
                self._toggle_cell_done_at(ev)
            return
        if self.base is None:
            return
        x, y = self._to_img(ev)
        if self.crop_box is None:
            self.crop_box = [x, y, None, None]
            self._draw_point_marker(x, y)
        else:
            self.crop_box[2], self.crop_box[3] = x, y
            self.show_crop_overlay()

    # ---------------------------------------------------------- 拼制进度跟随
    def _toggle_cell_done_at(self, ev):
        col = int((ev.x - self._img_x0) / self._cell_px) if self._cell_px else -1
        row = int((ev.y - self._img_y0) / self._cell_py) if self._cell_py else -1
        if not (0 <= row < self.grid_rows and 0 <= col < self.grid_cols):
            return
        key = (row, col)
        if key in self.done_cells:
            self.done_cells.discard(key)
        else:
            self.done_cells.add(key)
        self._redraw_done_marks()
        total = self.grid_cols * self.grid_rows
        done = len(self.done_cells)
        self.status("拼制进度：已拼 %d / %d 格（%.1f%%）——点击格子标记/取消，点「重置进度」清零"
                    % (done, total, done * 100.0 / total))

    def _redraw_done_marks(self):
        for it in self._done_items:
            try:
                self.canvas.delete(it)
            except Exception:
                pass
        self._done_items = []
        if not (self.grid_cols and self.done_cells and self._cell_px):
            return
        for (r, c) in self.done_cells:
            x0 = self._img_x0 + c * self._cell_px
            y0 = self._img_y0 + r * self._cell_py
            x1 = x0 + self._cell_px
            y1 = y0 + self._cell_py
            if (self._cell_px < 5 and self._cell_px > 0):
                # 太小看不清，仅显示中心小圆点
                self._done_items.append(
                    self.canvas.create_oval(x0 + 2, y0 + 2, x1 - 2, y1 - 2,
                                            fill="#1C1C1C", outline=""))
                continue
            e = max(2.0, min(self._cell_px, self._cell_py) * 0.16)
            # 深色粗描边 + 浅色内核，形成清晰的 X 记号
            self._done_items.append(self.canvas.create_line(
                x0 + e, y0 + e, x1 - e, y1 - e, fill="#1C1C1C", width=3))
            self._done_items.append(self.canvas.create_line(
                x0 + e, y1 - e, x1 - e, y0 + e, fill="#1C1C1C", width=3))
            self._done_items.append(self.canvas.create_line(
                x0 + e + 1, y0 + e + 1, x1 - e - 1, y1 - e - 1,
                fill="#FFFFFF", width=1))
            self._done_items.append(self.canvas.create_line(
                x0 + e + 1, y1 - e - 1, x1 - e - 1, y0 + e + 1,
                fill="#FFFFFF", width=1))

    def reset_progress(self):
        self.done_cells.clear()
        self._redraw_done_marks()
        self.status(L.tr("已清零拼制进度，重新开始标记"))

    def _draw_point_marker(self, x, y):
        if self.overlay_id is not None:
            try:
                self.canvas.delete(self.overlay_id)
            except Exception:
                pass
            self.overlay_id = None
        cx = 4 + x * self.canvas_scalex
        cy = 4 + y * self.canvas_scaley
        self.overlay_id = self.canvas.create_rectangle(cx - 3, cy - 3,
                                                       cx + 3, cy + 3,
                                                       outline="#ff3b30", width=2)

    def on_canvas_drag(self, ev):
        if (self.mode.get() != "manual" or self.crop_box is None
                or self.crop_box[2] is None):
            return
        x, y = self._to_img(ev)
        self.crop_box[2], self.crop_box[3] = x, y
        self.show_crop_overlay()

    def show_crop_overlay(self):
        if self.overlay_id is not None:
            try:
                self.canvas.delete(self.overlay_id)
            except Exception:
                pass
            self.overlay_id = None
        b = self.crop_box
        if not b or any(v is None for v in b):
            return
        x0 = 4 + b[0] * self.canvas_scalex
        y0 = 4 + b[1] * self.canvas_scaley
        x1 = 4 + b[2] * self.canvas_scalex
        y1 = 4 + b[3] * self.canvas_scaley
        self.overlay_id = self.canvas.create_rectangle(x0, y0, x1, y1,
                                                       outline="#ff3b30",
                                                       width=2, dash=(6, 3))

    def clear_crop(self):
        if self.crop_box:
            self.crop_box = None
        if self.overlay_id is not None:
            try:
                self.canvas.delete(self.overlay_id)
            except Exception:
                pass
            self.overlay_id = None

    def set_grid(self, w, h):
        self.lW.set(w)
        self.lH.set(h)

    # ---------------------------------------------------------- 推荐尺寸
    def _suggest_options(self):
        if self.base is None:
            return None
        try:
            cols, rows, base = self._grid_size()
        except Exception:
            cols, rows, base = int(self.lW.get()), int(self.lH.get()), self.base
        opts = be.suggest_sizes(base.width, base.height)
        return (cols, rows), base, opts

    def suggest_dialog(self, auto=False):
        info = self._suggest_options()
        if not info:
            return
        (cur_w, cur_h), base, opts = info
        if not opts:
            return
        dlg = tk.Toplevel(self)
        dlg.configure(bg=BG)
        dlg.title(L.tr("推荐图纸尺寸（按图片比例）"))
        dlg.transient(self)
        dlg.resizable(False, False)
        tk.Label(dlg, text="图片实际比例 %d × %d，按等比例推荐了下面几档尺寸："
                          % (base.width, base.height),
                 bg=BG, fg=INK, font=(FONT, 10)).grid(row=0, column=0,
                                                      columnspan=2, sticky=tk.W,
                                                      padx=16, pady=(14, 4))
        tk.Label(dlg, text=L.tr("拼豆数量越少越省事，越多越还原照片细节。"),
                 bg=BG, fg=INK_FAINT, font=(FONT, 8)).grid(row=1, column=0,
                                                           columnspan=2,
                                                           sticky=tk.W,
                                                           padx=16, pady=(0, 8))
        var = tk.StringVar(value="custom")

        def tier(total):
            if total <= 1500:
                return L.tr("小图 · 省豆快拼")
            if total <= 4000:
                return L.tr("中图 · 均衡")
            if total <= 10000:
                return L.tr("大图 · 细节丰富")
            if total <= 20000:
                return L.tr("超大图 · 高度还原")
            return L.tr("巨幅 · 精细还原")

        for i, (cols, rows, total) in enumerate(opts):
            d = abs(cols * rows - cur_w * cur_h)
            tag = L.tr("（当前") if d == 0 and cur_w and cur_h else ""
            ttk.Radiobutton(dlg, text="%d × %d　约 %d 颗　%s%s"
                            % (cols, rows, total, tag, tier(total)),
                            value=str(i), variable=var).grid(
                                row=2 + i, column=0, columnspan=2, sticky=tk.W,
                                padx=20, pady=2)
        ttk.Radiobutton(dlg, text="不动，保持当前尺寸（%d × %d）" % (cur_w, cur_h),
                        value="custom", variable=var).grid(
                            row=2 + len(opts), column=0, columnspan=2, sticky=tk.W,
                            padx=20, pady=(8, 12))
        if opts:
            close_i = min(range(len(opts)),
                          key=lambda i: abs(opts[i][0] * opts[i][1] - cur_w * cur_h))
            var.set(str(close_i))
        else:
            var.set("custom")
        btns = tk.Frame(dlg, bg=BG)
        btns.grid(row=2 + len(opts) + 1, column=0, columnspan=2, sticky=tk.E,
                  padx=16, pady=(0, 14))

        def apply():
            v = var.get()
            if v != "custom":
                cols, rows, _ = opts[int(v)]
                self.set_grid(cols, rows)
                self.status("已应用推荐尺寸 %d×%d" % (cols, rows))
            dlg.destroy()

        _mk_button(btns, L.tr("应 用"), apply).pack(side=tk.LEFT)
        Chip(btns, L.tr("取消"), dlg.destroy, color=CARD, fg=INK_SOFT).pack(
            side=tk.LEFT, padx=(8, 0))
        dlg.grab_set()
        dlg.wait_window()

    # ---------------------------------------------------------- 生成
    def _grid_size(self):
        cols = int(self.lW.get())
        rows = int(self.lH.get())
        if not (2 <= cols <= 400 and 2 <= rows <= 400):
            raise ValueError(L.tr("网格宽/高需在 2~400 之间"))
        base = self.base.copy()
        if self.crop_box and all(v is not None for v in self.crop_box):
            xs = sorted((self.crop_box[0], self.crop_box[2]))
            ys = sorted((self.crop_box[1], self.crop_box[3]))
            x0, x1, y0, y1 = xs[0], xs[1], ys[0], ys[1]
            if x1 - x0 >= 2 and y1 - y0 >= 2:
                base = base.crop((x0, y0, x1 + 1, y1 + 1))
        if self.keep_ratio.get():
            rows = max(2, round(cols * base.height / base.width))
        return cols, rows, base

    def _detect_subject(self):
        """检测主体并显示结果。"""
        if self.base is None:
            messagebox.showinfo(APP_NAME, L.tr("请先打开一张图片。"))
            return
        # 用当前抠图模式得到 base
        self.apply_cutout()
        _, subject_found, ratio = be.cartoonize(
            self.base, subject_detect=True, edge_size=1, smooth_level=1)
        if subject_found:
            self.cartoon_status.config(
                text="OK 主体已检测（%.0f%%）" % (ratio * 100), fg=TEAL)
        else:
            self.cartoon_status.config(
                text="主体不明显（%.0f%%），建议手动框选" % (ratio * 100), fg=BEAD)

    def _save_ai_key(self):
        """保存 SenseNova API key 到本地设置。"""
        key = self._ai_key_var.get().strip()
        if key:
            self._remember_settings()
            self._ai_status.config(text=L.tr("Key 已保存"), fg=TEAL)
        else:
            self._ai_status.config(text=L.tr("请输入 Key"), fg=BEAD)

    def _open_ai_guide(self, ev=None):
        """打开 AI 卡通 Key 获取教程页（本地 HTML，用系统浏览器）。"""
        try:
            path = _guide_path()
            if os.path.exists(path):
                webbrowser.open("file://" + os.path.abspath(path))
            else:
                webbrowser.open("https://platform.sensenova.cn/console/keys")
        except Exception as e:
            messagebox.showwarning(APP_NAME, "无法打开教程页：%s" % str(e))

    def _refresh_member_display(self):
        """刷新会员状态文字与颜色。"""
        if member_mod.is_active():
            self._member_status_var.set("✅ 会员有效期至 %s" %
                                        member_mod._fmt(member_mod.expire_at()))
            self._member_status_label.config(fg="#27AE60")
        elif member_mod.is_expired():
            self._member_status_var.set(L.tr("⚠ 会员已过期"))
            self._member_status_label.config(fg="#C0392B")
        else:
            self._member_status_var.set(L.tr("未开通会员"))
            self._member_status_label.config(fg="#8A93A0")

    def _open_member_dialog(self):
        """弹出会员激活对话框：收款码 + 兑换码输入。"""
        dlg = tk.Toplevel(self)
        dlg.title(L.tr("💎 开通 / 激活会员"))
        dlg.transient(self)
        dlg.geometry("360x520")
        dlg.resizable(False, False)
        dlg.configure(bg=BG)
        # 收款码
        # ★ 修复：_img 必须挂在对象上（dlg._pay_img），否则函数一返回就被
        #   垃圾回收，Tk 里的图片随之删除，Label 变空白 —— 这就是之前
        #   「弹窗里看不到收款码」的原因。
        try:
            _mp = os.path.join(os.path.dirname(_guide_path()), "payment_alipay.png")
            if os.path.exists(_mp):
                _pil = Image.open(_mp).convert("RGB")
                _w, _h = _pil.size
                _ratio = min(280 / _w, 320 / _h, 1.0)
                if _ratio < 1.0:
                    _pil = _pil.resize((max(1, int(_w * _ratio)),
                                        max(1, int(_h * _ratio))), Image.LANCZOS)
                dlg._pay_img = ImageTk.PhotoImage(_pil)
                tk.Label(dlg, image=dlg._pay_img, bg=BG).pack(pady=(16, 8))
            else:
                tk.Label(dlg, text=L.tr("（收款码图片待配置）"), bg=BG,
                         fg=INK_FAINT, font=(FONT, 9)).pack(pady=(16, 8))
        except Exception:
            tk.Label(dlg, text=L.tr("（收款码加载失败）"), bg=BG, fg="#C0392B",
                     font=(FONT, 9)).pack(pady=(16, 8))
        tk.Label(dlg, text=L.tr("支付宝扫码付款（9.9 元/月 · 29 元/年）"),
                 bg=BG, fg="#B9842C", font=(FONT, 10, "bold")).pack()
        tk.Label(dlg, text=L.tr("付款后向客服获取兑换码，输入下方激活"),
                 bg=BG, fg=INK_FAINT, font=(FONT, 8)).pack(pady=(4, 0))
        # 兑换码输入
        row = tk.Frame(dlg, bg=BG); row.pack(pady=(12, 4))
        tk.Label(row, text=L.tr("兑换码"), bg=BG, fg=INK,
                 font=(FONT, 9, "bold")).pack(side=tk.LEFT)
        code_var = tk.StringVar(value="")
        entry = ttk.Entry(row, textvariable=code_var, width=22,
                          font=(FONT, 11))
        entry.pack(side=tk.LEFT, padx=(8, 4), fill=tk.X, expand=True)
        entry.focus_set()
        # 激活按钮
        def _do_activate():
            code = code_var.get().strip().upper()
            if not code:
                messagebox.showwarning(L.tr("激活"), L.tr("请输入兑换码。"), parent=dlg)
                return
            try:
                member_mod.mark_code_used  # noqa: B018  存在性检查
            except Exception:
                pass
            ok, r = member_mod.verify_offline(code)
            if ok and r:
                member_mod.set_expire_at(r)
                member_mod.mark_code_used(code)
                self._refresh_member_display()
                dlg.destroy()
                messagebox.showinfo(L.tr("激活成功"),
                                     L.tr("🎉 会员激活成功！\n有效期至 %s") %
                                     member_mod._fmt(r), parent=self)
            else:
                messagebox.showerror(L.tr("激活失败"), r or L.tr("兑换码无效。"), parent=dlg)
        tk.Button(row, text=L.tr("激活"), command=_do_activate,
                  bg="#27AE60", fg="#FFFFFF", font=(FONT, 9, "bold"),
                  relief="flat", cursor="hand2", padx=12, pady=2).pack(
                      side=tk.LEFT, padx=(4, 0))
        # Enter 键激活
        entry.bind("<Return>", lambda e: _do_activate())

    def _activate_member(self):
        """兑换码激活（本地离线校验）。"""
        code = self._member_code_var.get().strip().upper()
        if not code:
            messagebox.showwarning(APP_NAME, L.tr("请输入兑换码。"))
            return
        self._activate_btn.config(state="disabled")
        self.status(L.tr("正在激活会员…"))
        threading.Thread(target=self._activate_member_worker,
                         args=(code, member_mod.get_device_id()),
                         daemon=True).start()

    def _activate_member_worker(self, code, device_id):
        """后台线程：本地离线校验兑换码。"""
        ok, result = member_mod.verify_offline(code)
        expire_at = result if ok else None
        err_msg = result if not ok else ""
        self.after(0, lambda: self._activate_member_done(ok, err_msg, expire_at))

    def _activate_member_done(self, ok, err_msg, expire_at):
        self._activate_btn.config(state="normal")
        if ok and expire_at:
            code = self._member_code_var.get().strip().upper()
            member_mod.set_expire_at(expire_at)
            member_mod.mark_code_used(code)
            self._refresh_member_display()
            self.status("会员激活成功，有效期至 %s" % member_mod._fmt(expire_at))
            messagebox.showinfo(APP_NAME, "🎉 会员激活成功！\n有效期至 %s" % member_mod._fmt(expire_at))
        else:
            self.status("激活失败：%s" % err_msg)
            messagebox.showerror(APP_NAME, err_msg)

    def _ai_cartoon(self):
        """AI 卡通化入口：校验后开线程调 API，成功后替换原图。
        等待期间显示阶段文字 + 转动符号（非百分比进度，接口不返回进度）。"""
        if self._ai_cartoon_running:
            return
        if self.base is None:
            messagebox.showinfo(APP_NAME, L.tr("请先打开一张图片。"))
            return
        key = self._ai_key_var.get().strip()
        if not key:
            self._ai_status.config(text=L.tr("请先填写 API Key"), fg=BEAD)
            messagebox.showwarning(APP_NAME, "请先在上方填写 SenseNova API Key。\n在 platform.sensenova.cn 获取。")
            return
        prompt = self._ai_prompt_var.get().strip() or L.tr("把它变成可爱卡通风格，Q版动漫，色彩明亮")
        self._ai_cartoon_running = True
        self._ai_spinner_idx = 0
        self._ai_phase = L.tr("准备中…")
        self._ai_btn.set_state(True)
        self._ai_btn.set_text(L.tr("生成中…"))
        self._ai_status.config(text=L.tr("⏳ 准备中…"), fg="#B9842C")
        self.status(L.tr("AI 卡通化中…"))
        self.update_idletasks()
        threading.Thread(
            target=self._ai_cartoon_worker, args=(self.base, key, prompt),
            daemon=True).start()
        self._ai_spinner_tick()

    def _ai_spinner_tick(self):
        """转动符号循环（⏳/◐/◑/◒），仅表示「进行中」，不代表真实进度百分比。"""
        if not self._ai_cartoon_running:
            return
        spins = ("◐", "◑", "◒", "❒")
        phase = getattr(self, "_ai_phase", L.tr("处理中"))
        self._ai_status.config(text="%s  %s" % (spins[self._ai_spinner_idx % len(spins)], phase),
                               fg="#B9842C")
        self._ai_spinner_idx += 1
        self.status("AI 卡通化中… %s" % phase)
        self.after(450, self._ai_spinner_tick)

    def _ai_cartoon_worker(self, image, key, prompt):
        try:
            self.after(0, lambda: setattr(self, "_ai_phase", L.tr("正在生成，请稍候（约数十秒）")))
            result = be.ai_cartoonize(image, api_key=key, prompt=prompt)
        except be.AICartoonError as e:
            err = str(e)
            result = None
        except Exception as e:
            err = "AI 卡通生成失败：%s" % str(e)
            result = None
        else:
            err = None
        self.after(0, lambda: self._ai_cartoon_done(result, err))

    def _ai_cartoon_done(self, result, err):
        self._ai_cartoon_running = False
        self._ai_btn.set_state(False)
        self._ai_btn.set_text(L.tr("AI 卡通化"))
        if err:
            self._ai_status.config(text=L.tr("✕ 失败"), fg=BEAD)
            self.status(L.tr("AI 卡通化失败"))
            messagebox.showerror(APP_NAME, err)
            return
        self._ai_status.config(text=L.tr("✔ OK 已替换原图"), fg=TEAL)
        self.base = result.convert("RGBA")
        self.show_pil(self.base)
        self.status(L.tr("AI 卡通化完成，已替换原图；可点击「生成图纸」"))
        # 若已有图纸，用卡通图重算预览（可选：让用户自己点生成更明确，此处仅刷新原图）

    def generate(self):
        if self.base is None:

            messagebox.showinfo(APP_NAME, L.tr("请先打开一张图片。"))
            return

        self.prog.reset()
        self.prog.pack(fill=tk.X, pady=(0, 8))
        self.gen_btn.set_state(True)
        self.gen_btn.set_text(L.tr("生成中…"))
        self.update_idletasks()

        try:
            cols, rows, base = self._grid_size()
            # 阶段 0：卡通化（可选）
            if self.cartoon_enabled.get():
                edge = int(self.cartoon_edge.get())
                smooth = int(self.cartoon_smooth.get())
                self.status(L.tr("正在卡通化…"))
                self.update_idletasks()
                cartoon, subject_found, subject_ratio = be.cartoonize(
                    base, subject_detect=True, edge_size=edge, smooth_level=smooth)
                if not subject_found:
                    messagebox.showwarning(
                        APP_NAME,
                        "未检测到明显主体（主体只占画面 %.0f%%）。\n"
                        "建议使用「手动框选」圈出主体，或换一张主体清晰的照片。"
                        % (subject_ratio * 100))
                base = cartoon.convert("RGBA")
            self._remember_settings()   # 生成即记住当前参数
            # 阶段 1：网格化
            self.prog.set_value(0, 0.0, label=self.prog.STAGES[0])
            self.status(L.tr("正在网格化图片…"))
            self.update_idletasks()
            cells, bg_hex = be.grid_by_average(base, cols, rows,
                                               keep_alpha=self.remove_alpha)
            self.cached_cells = cells
            self.prog.set_done(0)

            # 阶段 2：颜色映射
            palette = self._find_palette(self.pal_var.get())
            if palette is None:   # 所选色卡已被删除等异常情况
                palette = palettes.get_all_palettes()[0]
                self._refresh_palette_box(palette["name"])
            bg_name = "白色" if self.bg_fill.get() else None
            self.prog.set_value(1, 0.0, label=self.prog.STAGES[1])
            self.status(L.tr("正在映射颜色到色卡…"))
            self.update_idletasks()
            result, used = be.map_to_palette(cells, palette["colors"],
                                             int(self.maxC.get()),
                                             bg_overwrite=bg_name, image=base)
            total = sum(c for _, _, c in used)
            self.result, self.used_colors = result, used
            self.grid_cols, self.grid_rows = cols, rows
            self.prog.set_done(1)

            # 阶段 3：渲染预览
            self.prog.set_value(2, 0.0, label=self.prog.STAGES[2])
            self.status(L.tr("正在渲染预览…"))
            self.update_idletasks()
            self.highlight_color = None
            self.done_cells = set()   # 新图纸，进度清零
            self._done_items = []
            self.color_preview = be.render_grid_preview(cells, result, cols, rows,
                                                        cell_px=120, out_w=760)
            self.number_preview = be.render_number_view(result, cols, rows, used,
                                                        cell_px=36, margin=16)
            self.showing_number = False
            self.show_pil_right(self.color_preview)
            # 单色分布条
            self._build_swatches([(n, hx) for n, hx, _c in used])
            self._swatch_open = False
            self.swatch_canvas.pack_forget()
            self.swatch_arrow.config(text="▶")
            self.prog.set_done(2)
            self.status(f"{L.tr('生成完成：')}{cols}x{rows}，{len(used)}{L.tr(' 种颜色，共 ')}{total}{L.tr(' 颗豆')}")
            lines_txt = [f"{L.tr('图纸尺寸：')}{cols} x {rows}{L.tr('   用色：')}{len(used)}{L.tr(' 种   共需：')}{total}{L.tr(' 颗豆')}",
                         "色卡：%s   模式：%s" % (self.pal_var.get(),
                           self.mode.get().replace("auto", L.tr("自动去背景"))
                           .replace("manual", L.tr("手动框选"))
                           .replace("none", L.tr("不抠图")))]
            if used:
                top = used[:3]
                lines_txt.append(L.tr("用豆最多：") +
                                 "、".join(f"{n}({c})" for n, _, c in top))
            self.stats_var.set("\n".join(lines_txt))
        except Exception as e:
            self.prog.reset()
            messagebox.showerror(APP_NAME, f"{L.tr('生成失败：')}{e}{L.tr('\\n\\n请调整参数后重试。')}")
        finally:
            self.gen_btn.set_state(False)
            self.gen_btn.set_text(L.tr("生 成 图 纸"))

    def toggle_view(self):
        if self.result is None:
            return
        self.showing_number = not self.showing_number
        if self.showing_number:
            self.show_pil_right(self.number_preview)
            self.toggle_btn.set_text(L.tr("◀ 颜色图"))
            self.status(L.tr("编号图模式：格内数字=颜色编号，底部图例对照"))
        else:
            self.show_pil_right(self.color_preview)
            self.toggle_btn.set_text("编号图 ▶")
            self.status(L.tr("颜色图模式：直接看颜色拼豆"))

    # ---------------------------------------------------------- 镜像翻转
    def flip_view(self, axis):
        """水平/垂直镜像翻转图纸（翻转后颜色集合不变，直接重渲染预览）。"""
        if self.result is None:
            messagebox.showinfo(APP_NAME, L.tr("请先生成图纸，再使用翻转。"))
            return
        self.result = be.flip_pattern(self.result, self.grid_cols,
                                      self.grid_rows, axis)
        self.done_cells.clear()   # 格子位置变了，旧进度标记作废
        self._redraw_done_marks()
        # 重渲染两种预览（保留当前单色高亮）
        hl = self.highlight_color
        self.color_preview = be.render_grid_preview(
            self.cached_cells, self.result, self.grid_cols, self.grid_rows,
            cell_px=120, out_w=760, highlight=hl)
        self.number_preview = be.render_number_view(
            self.result, self.grid_cols, self.grid_rows, self.used_colors,
            cell_px=36, margin=16, highlight=hl)
        if self.showing_number:
            self.show_pil_right(self.number_preview)
        else:
            self.show_pil_right(self.color_preview)
        self.status("已%s翻转图纸，颜色分布已更新" % (L.tr("水平") if axis == "h" else L.tr("垂直")))

    # ---------------------------------------------------------- 单色分布视图
    def _toggle_swatches(self, ev=None):
        """展开/收起单色分布面板。"""
        if self._swatch_open:
            self.swatch_canvas.pack_forget()
            self.swatch_arrow.config(text="▶")
        else:
            self.swatch_canvas.pack(fill=tk.X, padx=6, pady=4)
            self.swatch_arrow.config(text="▼")
        self._swatch_open = not self._swatch_open

    def _on_swatch_wheel(self, ev):
        """单色条内滚轮 = 横向滚动（避免误触画布缩放）。"""
        if not self.swatch_inner.winfo_exists():
            return
        inner_w = self.swatch_inner.winfo_reqwidth()
        cw = self.swatch_canvas.winfo_width()
        if inner_w <= cw:
            return
        step = -1 if ev.delta > 0 else 1
        self.swatch_canvas.xview_scroll(2 * step, "units")
        return "break"

    def _build_swatches(self, colors):
        """colors: [(name, hex), ...]，构建可点击的色板条。"""
        for w in self.swatch_inner.winfo_children():
            w.destroy()
        self._swatch_chips = []

        def add_swatch(label, hx, name, active_fg="#FFFFFF"):
            cv = tk.Canvas(self.swatch_inner, width=78, height=40,
                           bg="#F5F7FA", highlightthickness=0, cursor="hand2")
            cv.grid(row=0, column=len(self._swatch_chips), padx=(0, 5))
            cv.create_rectangle(3, 4, 30, 36, fill=hx, outline="#8A93A0",
                                width=1)
            cv.create_text(36, 20, text=label, fill=INK, font=(FONT, 8),
                           anchor=tk.W)
            otl = cv.create_rectangle(1, 1, 77, 39, outline="",
                                      width=2)   # 选中描边（默认隐藏）
            cv.bind("<Button-1>", lambda e, nm=name: self._apply_highlight(nm))
            self._swatch_chips.append((name, cv, otl))

        add_swatch(L.tr("全部"), "#888891", None, None)
        for name, hx in colors:
            add_swatch(name, hx, name, None)

    def _refresh_swatch_styles(self):
        for name, cv, otl in self._swatch_chips:
            if name == self.highlight_color:
                cv.itemconfig(otl, outline=BEAD)
            else:
                cv.itemconfig(otl, outline="")

    def _apply_highlight(self, name):
        if name == self.highlight_color:
            name = None             # 再点同一颜色 = 恢复全部
        self.highlight_color = name
        self._refresh_swatch_styles()
        if self.result is None:
            return
        if self.showing_number:     # 编号图同样支持单色高亮
            self._rebuild_number_preview(highlight=name)
            self.show_pil_right(self.number_preview)
        else:
            self._rebuild_color_preview(highlight=name)
        if name:
            self.status("单色分布：只看「%s」的格子（其余淡显）" % name)
        else:
            self.status(L.tr("已恢复显示全部颜色"))

    def _rebuild_color_preview(self, highlight=None):
        self.color_preview = be.render_grid_preview(
            self.cached_cells, self.result, self.grid_cols, self.grid_rows,
            cell_px=120, out_w=760, highlight=highlight)
        self.show_pil_right(self.color_preview)

    def _rebuild_number_preview(self, highlight=None):
        self.number_preview = be.render_number_view(
            self.result, self.grid_cols, self.grid_rows, self.used_colors,
            cell_px=36, margin=16, highlight=highlight)
        self.show_pil_right(self.number_preview)

    # ---------------------------------------------------------- 导出（带预览）
    def _make_sheet(self):
        """生成用于保存的完整图纸 PNG（含标题、图例、编号图）。"""
        total = sum(c for _, _, c in self.used_colors)
        return be.render_pattern(
            self.cached_cells, self.result, self.used_colors,
            self.grid_cols, self.grid_rows, cell_px=26,
            title=f"{L.tr('拼豆图纸  ')}{self.grid_cols}×{self.grid_rows} · "
                  f"{len(self.used_colors)}{L.tr('色 · 共')}{total}{L.tr('颗')}")

    def preview_pattern_dialog(self, sheet=None):
        """保存图纸前弹窗：预览图纸缩略图 + 确认保存。"""
        if sheet is None:
            sheet = self._make_sheet()
        dlg = tk.Toplevel(self)
        dlg.configure(bg=BG)
        dlg.title(L.tr("预览图纸 · 确认保存"))
        dlg.transient(self)
        dlg.resizable(False, False)
        tk.Label(dlg, text=L.tr("下面是将要保存的完整图纸（可打印）："),
                 bg=BG, fg=INK, font=(FONT, 10, "bold")).pack(
                     anchor=tk.W, padx=18, pady=(14, 6))
        # 缩略预览
        th = sheet.copy()
        th.thumbnail((640, 720))
        photo = ImageTk.PhotoImage(th)
        cv = tk.Canvas(dlg, width=th.width + 8, height=th.height + 8,
                      bg=PAPER, highlightthickness=1,
                      highlightbackground="#C4CBD6")
        cv.pack(padx=18, pady=(0, 6))
        cv.create_image(4, 4, anchor=tk.NW, image=photo)
        btns = tk.Frame(dlg, bg=BG)
        btns.pack(pady=(0, 14))
        _mk_button(btns, L.tr("保存图纸"), lambda: self._do_save_pattern(dlg, sheet)).pack(
            side=tk.LEFT)
        Chip(btns, L.tr("取消"), dlg.destroy, color=CARD, fg=INK_SOFT).pack(
            side=tk.LEFT, padx=(8, 0))
        dlg.photo_ref = photo
        dlg.grab_set()
        dlg.wait_window()

    def _do_save_pattern(self, dlg, sheet):
        dlg.destroy()
        path = filedialog.asksaveasfilename(
            title=L.tr("保存拼豆图纸"), defaultextension=".png",
            initialfile=f"{L.tr('拼豆图纸_')}{self.grid_cols}x{self.grid_rows}.png",
            filetypes=[(L.tr("PNG 图片"), "*.png"), (L.tr("所有文件"), "*.*")])
        if not path:
            return
        try:
            sheet.save(path, dpi=(150, 150))
            self.status(f"{L.tr('图纸已保存：')}{path}")
        except Exception as e:
            messagebox.showerror(APP_NAME, f"{L.tr('保存失败：')}{e}")

    def save_pattern(self):
        if self.result is None:
            messagebox.showinfo(APP_NAME, L.tr("还没有可保存的图纸，先点「生成图纸」。"))
            return
        try:
            sheet = self._make_sheet()
        except Exception as e:
            messagebox.showerror(APP_NAME, f"{L.tr('生成图纸失败：')}{e}")
            return
        self.preview_pattern_dialog(sheet)

    def _list_lines(self):
        total = sum(c for _, _, c in self.used_colors)
        lines = [f"{L.tr('拼豆耗材清单\u3000图纸 ')}{self.grid_cols}×{self.grid_rows}",
                 f"{L.tr('用色 ')}{len(self.used_colors)}{L.tr(' 种，共 ')}{total}{L.tr(' 颗\\n')}"]
        for i, (name, hx, cnt) in enumerate(self.used_colors, 1):
            lines.append(f"{i:>3}. {name:<7} #{hx.lstrip('#'):<7}{L.tr('  需要 ')}{cnt:>5}{L.tr(' 颗')}")
        lines.append("\n提示：留空格已跳过，不占用耗材。")
        return lines

    def preview_list_dialog(self):
        """保存耗材清单前弹窗：文本预览 + 确认保存。"""
        text = "\n".join(self._list_lines())

        dlg = tk.Toplevel(self)
        dlg.configure(bg=BG)
        dlg.title(L.tr("预览耗材清单 · 确认保存"))
        dlg.transient(self)
        dlg.resizable(False, False)
        tk.Label(dlg, text=L.tr("下面是将要保存的耗材清单（txt）："),
                 bg=BG, fg=INK, font=(FONT, 10, "bold")).pack(
                     anchor=tk.W, padx=18, pady=(14, 6))
        txt = tk.Text(dlg, width=46, height=16, font=(FONT, 10),
                      bg=PAPER, fg=INK, relief="flat", borderwidth=0)
        txt.pack(padx=18, pady=(0, 6))
        txt.insert("1.0", text)
        txt.config(state="disabled")
        btns = tk.Frame(dlg, bg=BG)
        btns.pack(pady=(0, 14))
        _mk_button(btns, L.tr("保存清单"), lambda: self._do_save_list(dlg)).pack(
            side=tk.LEFT)
        Chip(btns, L.tr("另存为 Excel"), lambda: self.export_list_excel(dlg),
             color=TEAL, fg="#FFFFFF", font=(FONT, 9, "bold")).pack(
                 side=tk.LEFT, padx=(8, 0))
        Chip(btns, L.tr("取消"), dlg.destroy, color=CARD, fg=INK_SOFT).pack(
            side=tk.LEFT, padx=(8, 0))
        dlg.grab_set()
        dlg.wait_window()

    def _do_save_list(self, dlg):
        dlg.destroy()
        path = filedialog.asksaveasfilename(
            title="保存耗材清单", defaultextension=".txt",
            initialfile=L.tr("拼豆耗材清单.txt"),
            filetypes=[(L.tr("文本文件"), "*.txt"), (L.tr("所有文件"), "*.*")])
        if not path:
            return
        try:
            total = sum(c for _, _, c in self.used_colors)
            lines = [f"{L.tr('拼豆耗材清单\u3000图纸 ')}{self.grid_cols}×{self.grid_rows}",
                     f"{L.tr('用色 ')}{len(self.used_colors)}{L.tr(' 种，共 ')}{total}{L.tr(' 颗\\n')}"]
            for i, (name, hx, cnt) in enumerate(self.used_colors, 1):
                lines.append(f"{i:>2}. {name:<6} #{hx.lstrip('#'):<6}{L.tr('  需要 ')}{cnt:>5}{L.tr(' 颗')}")
            lines.append("\n提示：留空格已跳过，不占用耗材。")
            with open(path, "w", encoding="utf-8-sig") as f:
                f.write("\n".join(lines))
            self.status(f"{L.tr('耗材清单已保存：')}{path}")
        except Exception as e:
            messagebox.showerror(APP_NAME, f"{L.tr('保存失败：')}{e}")

    def save_list(self):
        if self.result is None:
            messagebox.showinfo(APP_NAME, L.tr("还没有生成结果。"))
            return
        self.preview_list_dialog()

    def export_list_excel(self, parent_dlg=None):
        """把耗材清单导出为 .xlsx（需要 openpyxl）。"""
        if self.result is None:
            messagebox.showinfo(APP_NAME, L.tr("还没有生成结果。"))
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showinfo(
                APP_NAME, "本机未安装 openpyxl：\n"
                "源码版请在命令行执行 pip install openpyxl 后重试。")
            return
        path = filedialog.asksaveasfilename(
            title=L.tr("导出耗材清单 Excel"), defaultextension=".xlsx",
            initialfile=L.tr("拼豆耗材清单.xlsx"),
            filetypes=[(L.tr("Excel 工作簿"), "*.xlsx"), (L.tr("所有文件"), "*.*")])
        if not path:
            return
        try:
            self._write_list_excel(path)
            self.status(f"{L.tr('耗材清单 Excel 已导出：')}{path}")
            if parent_dlg is not None:
                try:
                    parent_dlg.destroy()
                except Exception:
                    pass
        except Exception as e:
            messagebox.showerror(APP_NAME, f"{L.tr('导出 Excel 失败：')}{e}")

    def _write_list_excel(self, path):
        """生成 .xlsx 写盘（供 UI 按钮和自动化测试共用）。"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = L.tr("耗材清单")
        total = sum(c for _, _, c in self.used_colors)
        ws.append([L.tr("拼豆耗材清单"),
                   L.tr("图纸尺寸"), "%d × %d" % (self.grid_cols, self.grid_rows),
                   L.tr("用色"), len(self.used_colors), L.tr("共需"), total, L.tr("颗")])
        ws.append([])
        head = [L.tr("编号"), L.tr("颜色名"), L.tr("色值 (HEX)"), L.tr("所需颗数")]
        ws.append(head)
        hf = Font(bold=True, color="FFFFFF", size=11)
        hfill = PatternFill("solid", fgColor="2E8C83")
        for cell in ws[3]:
            cell.font = hf
            cell.fill = hfill
            cell.alignment = Alignment(horizontal="center")
        for i, (name, hx, cnt) in enumerate(self.used_colors, 1):
            ws.append([i, name, hx, cnt])
        for row in ws.iter_rows(min_row=4, min_col=3, max_col=3):
            for cell in row:
                try:
                    cell.fill = PatternFill("solid",
                                            fgColor=cell.value.lstrip("#"))
                except Exception:
                    pass
                cell.alignment = Alignment(horizontal="center")
        for col, wdt in (("A", 8), ("B", 16), ("C", 16), ("D", 12)):
            ws.column_dimensions[col].width = wdt
        wb.save(path)

    # ---------------------------------------------------------- PDF 打印导出
    @staticmethod
    def _split_pdf_pages(img, page_w=1240, page_h=1754, margin=48):
        """把一张长图拆成 A4(150dpi) 页面：能放下就单页居中，否则横向整页纵向切。"""
        box_w, box_h = page_w - 2 * margin, page_h - 2 * margin
        if img.width <= box_w and img.height <= box_h:
            page = Image.new("RGB", (page_w, page_h), "white")
            page.paste(img, ((page_w - img.width) // 2,
                             (page_h - img.height) // 2))
            return [page]
        scale = min(box_w / img.width, 3.0)
        w = max(1, int(img.width * scale))
        h = max(1, int(img.height * scale))
        scaled = img.resize((w, h), Image.LANCZOS)
        if scaled.width > box_w:   # 宽度超页：再等比缩小到页宽
            s2 = box_w / scaled.width
            scaled = scaled.resize((box_w, max(1, int(h * s2))),
                                   Image.LANCZOS)
        pages = max(1, -(-scaled.height // box_h))   # 向上取整页数
        out = []
        for i in range(pages):
            y0 = i * box_h
            y1 = min(scaled.height, (i + 1) * box_h)
            band = scaled.crop((0, y0, scaled.width, y1))
            page = Image.new("RGB", (page_w, page_h), "white")
            page.paste(band, ((page_w - band.width) // 2,
                              margin + (box_h - band.height) // 2))
            out.append(page)
        return out

    def _render_list_page(self):
        """把耗材清单渲染成一页 A4 图像。"""
        page = Image.new("RGB", (1240, 1754), "white")
        d = ImageDraw.Draw(page)
        font_title = be._load_font(30, bold=True)
        font = be._load_font(24)
        lines = self._list_lines()
        y = 90
        for i, ln in enumerate(lines):
            d.text((90, y), ln, font=font_title if i <= 1 else font,
                   fill="#1A1A1A")
            y += 52 if i > 1 else 68
        return page

    def _build_pdf_pages(self):
        sheet = self._make_sheet()
        pages = self._split_pdf_pages(sheet)
        pages.append(self._render_list_page())
        return pages

    def _write_pdf(self, path):
        pages = self._build_pdf_pages()
        pages[0].save(path, "PDF", save_all=True,
                      append_images=pages[1:], resolution=150.0)
        return len(pages)

    def export_pdf(self):
        """导出可打印 PDF：图纸多页 + 耗材清单页。"""
        if self.result is None:
            messagebox.showinfo(APP_NAME, L.tr("还没有可导出的图纸，先点「生成图纸」。"))
            return
        path = filedialog.asksaveasfilename(
            title=L.tr("导出 PDF（可打印）"), defaultextension=".pdf",
            initialfile=f"{L.tr('拼豆图纸_')}{self.grid_cols}x{self.grid_rows}.pdf",
            filetypes=[(L.tr("PDF 文档"), "*.pdf"), (L.tr("所有文件"), "*.*")])
        if not path:
            return
        try:
            self.status(L.tr("正在生成 PDF…"))
            self.update_idletasks()
            n = self._write_pdf(path)
            self.status(f"{L.tr('PDF 已导出：')}{path}{L.tr('（共 ')}{n}{L.tr(' 页）')}")
        except Exception as e:
            messagebox.showerror(APP_NAME, f"{L.tr('导出 PDF 失败：')}{e}")

    # ---------------------------------------------------------- 拼豆灵感
    def open_inspiration(self, page=0):
        """「拼豆灵感」面板：本地灵感库 + 在线找图，可一键用作底图。"""
        dlg = tk.Toplevel(self)
        dlg.configure(bg=BG)
        dlg.title(L.tr("拼豆灵感 · 挑一张好图开始拼"))
        dlg.transient(self)
        dlg.geometry("940x660")
        self._insp_dlg = dlg
        nb = ttk.Notebook(dlg)
        style = ttk.Style(dlg)
        for cl, bg in (("TNotebook", BG), ("TNotebook.Tab", BG)):
            try:
                style.configure(cl, background=bg)
            except Exception:
                pass
        nb.pack(fill=tk.BOTH, expand=True, padx=12, pady=(14, 6))
        self._insp_nb = nb
        tab1 = tk.Frame(nb, bg=BG)
        tab2 = tk.Frame(nb, bg=BG)
        nb.add(tab1, text=L.tr("  精选·本地灵感  "))
        nb.add(tab2, text=L.tr("  在线找灵感  "))
        nb.select(page)
        self._build_insp_tab(tab1)
        self._build_web_tab(tab2)
        dlg.grab_set()

    # ---------- 页1：本地灵感网格 ----------
    def _build_insp_tab(self, parent):
        self._insp_photos = []      # 缩略图引用
        self._insp_items = []       # [{name, path, source_url, keyword}]
        self._insp_idx = None       # 当前选中 index
        self._insp_focus = None     # 预览引用

        left = tk.Frame(parent, bg=BG)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(8, 4), pady=8)
        if not ins.scan():
            tk.Label(left, text="灵感库还是空的\n点下面「在线找灵感」检索，或「从本地导入图片」",
                     bg=PAPER, fg=INK_FAINT, font=(FONT, 10)).pack(pady=30)
        wrap = tk.Frame(left, bg=PAPER, bd=1, relief="solid")
        wrap.pack(fill=tk.BOTH, expand=True)
        self._insp_canvas = tk.Canvas(wrap, bg=PAPER, highlightthickness=0)
        sb = tk.Scrollbar(wrap, orient=tk.VERTICAL, command=self._insp_canvas.yview)
        self._insp_canvas.configure(yscrollcommand=sb.set)
        self._insp_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self._insp_grid = tk.Frame(self._insp_canvas, bg=PAPER)
        self._insp_canvas.create_window((0, 0), window=self._insp_grid,
                                        anchor=tk.NW)
        self._insp_grid.bind(
            "<Configure>",
            lambda e: self._insp_canvas.configure(
                scrollregion=self._insp_canvas.bbox("all")))
        self._insp_canvas.bind(
            "<MouseWheel>",
            lambda e: self._insp_canvas.yview_scroll(-e.delta // 120, "units"))

        right = tk.Frame(parent, bg=BG, width=290)
        right.pack(side=tk.RIGHT, fill=tk.Y, padx=(4, 8), pady=8)
        right.pack_propagate(False)
        tk.Label(right, text=L.tr("预览"), bg=BG, fg=INK, font=(FONT, 10, "bold")
                 ).pack(anchor=tk.W, pady=(0, 4))
        prev_card = tk.Frame(right, bg=PAPER, bd=1, relief="solid")
        prev_card.pack(fill=tk.BOTH, expand=True)
        self._insp_preview = tk.Label(prev_card, bg=PAPER, text=L.tr("点选左侧一张图"),
                                      fg=INK_FAINT, font=(FONT, 9))
        self._insp_preview.pack(fill=tk.BOTH, expand=True)
        self._insp_meta = tk.Label(right, text="", bg=BG, fg=INK_SOFT,
                                   font=(FONT, 8), justify=tk.LEFT,
                                   wraplength=280, anchor=tk.W)
        self._insp_meta.pack(fill=tk.X, pady=(6, 4))
        tk.Frame(right, bg=BG, height=6).pack()
        _mk_button(right, L.tr("用作底图"), self._insp_use_as_base).pack(
            fill=tk.X, pady=(2, 6))
        Chip(right, L.tr("从本地导入图片"), self._insp_import, color=TEAL,
             fg="#FFFFFF", font=(FONT, 9, "bold")).pack(fill=tk.X)
        tk.Label(right, text=L.tr("提示：选中后点「用作底图」即可进入抠图 / 尺寸 / 色卡 / 生成流程"),
                 bg=BG, fg=INK_FAINT, font=(FONT, 8), justify=tk.LEFT,
                 wraplength=280).pack(anchor=tk.W, pady=(8, 0))
        self._insp_refresh()

    def _insp_refresh(self):
        for w in self._insp_grid.winfo_children():
            w.destroy()
        self._insp_photos = []
        self._insp_items = ins.scan()
        self._insp_idx = None
        ncols = 4
        for i, item in enumerate(self._insp_items):
            row, col = i // ncols, i % ncols
            card = tk.Frame(self._insp_grid, bg=PAPER, bd=1, relief="solid",
                            cursor="hand2")
            card.grid(row=row, column=col, padx=6, pady=6)
            cv = tk.Canvas(card, width=104, height=104, bg=PAPER,
                           highlightthickness=0)
            cv.pack()
            photo = self._thumb_photo(item["path"])
            if photo:
                cv.create_image(52, 52, image=photo)
                self._insp_photos.append(photo)
            tk.Label(card, text=item["name"], bg=PAPER, fg=INK,
                     font=(FONT, 8)).pack(pady=(0, 4))
            card.bind("<Button-1>", lambda e, k=i: self._insp_select(k))
            cv.bind("<Button-1>", lambda e, k=i: self._insp_select(k))
            card.bind("<Double-Button-1>", lambda e, k=i: self._insp_open(k))
            cv.bind("<Double-Button-1>", lambda e, k=i: self._insp_open(k))

    def _thumb_photo(self, path):
        try:
            im = Image.open(path).convert("RGB")
            im.thumbnail((96, 96))
            return ImageTk.PhotoImage(im)
        except Exception:
            return None

    def _insp_select(self, idx):
        self._insp_idx = idx
        item = self._insp_items[idx]
        try:
            im = Image.open(item["path"]).convert("RGB")
            im.thumbnail((270, 270))
            ph = ImageTk.PhotoImage(im)
        except Exception:
            ph = None
        if ph:
            self._insp_photos.append(ph)
            self._insp_preview.configure(image=ph, text="")
        else:
            self._insp_preview.configure(image="", text=L.tr("无法预览"))
        meta = "名称：%s\n来源：%s\n关键词：%s" % (
            item["name"], item["source_url"] or L.tr("本地/程序生成"),
            item["keyword"] or "—")
        self._insp_meta.configure(text=meta)

    def _insp_open(self, idx):
        self._insp_select(idx)
        self._insp_use_as_base()

    def _insp_use_as_base(self):
        if self._insp_idx is None:
            messagebox.showinfo(APP_NAME, L.tr("先在左侧选中一张灵感图。"))
            return
        item = self._insp_items[self._insp_idx]
        try:
            self.src_img = be.load_image(item["path"])
        except Exception as e:
            messagebox.showerror(APP_NAME, f"{L.tr('加载灵感图失败：')}{e}")
            return
        self._last_path = item["path"]
        self.crop_box = None
        self.result = None
        self.highlight_color = None
        self.mode.set("none")     # 灵感图多为图纸/成品图，默认整张使用
        self.apply_cutout(reinteractive=True)
        try:
            self.swatch_canvas.pack_forget()
            self._swatch_open = False
            self.swatch_arrow.config(text="▶")
        except Exception:
            pass
        try:
            self._insp_dlg.destroy()
        except Exception:
            pass
        self.status("灵感图已载入：「%s」—— 调整尺寸/色卡后点「生成图纸」" % item["name"])

    def _insp_import(self):
        paths = filedialog.askopenfilenames(
            title=L.tr("导入拼豆灵感图（复制到灵感库）"),
            filetypes=[(L.tr("图片文件"), "*.png *.jpg *.jpeg *.bmp *.webp"),
                       (L.tr("所有文件"), "*.*")])
        if not paths:
            return
        d = ins.inspiration_dir()
        os.makedirs(d, exist_ok=True)
        n = 0
        for i, p in enumerate(paths):
            try:
                base = "自定义_%02d%s" % (i + 1,
                                          os.path.splitext(p)[1].lower() or ".png")
                dst = os.path.join(d, base)
                shutil.copy2(p, dst)
                ins.record_source(d, base, "", L.tr("本地导入"))
                n += 1
            except Exception:
                continue
        self._insp_refresh()
        if n:
            self.status("已导入 %d 张灵感图到本地灵感库。" % n)

    # ---------- 页2：在线找灵感 ----------
    def _build_web_tab(self, parent):
        tk.Label(parent, text=L.tr("在线检索公开图源（Bing 图片索引，含小红书/堆糖等平台的拼豆图）"),
                 bg=BG, fg=INK, font=(FONT, 11, "bold")).pack(
                     anchor=tk.W, padx=14, pady=(14, 4))
        row = tk.Frame(parent, bg=BG)
        row.pack(anchor=tk.W, padx=14, pady=(0, 4))
        tk.Label(row, text=L.tr("关键词"), bg=BG, fg=INK, font=(FONT, 9)).pack(side=tk.LEFT)
        self._web_kw = tk.Entry(row, width=22, font=(FONT, 10))
        self._web_kw.insert(0, "拼豆图纸")
        self._web_kw.pack(side=tk.LEFT, padx=6)
        tk.Label(row, text=L.tr("数量"), bg=BG, fg=INK, font=(FONT, 9)).pack(side=tk.LEFT)
        self._web_n = ttk.Spinbox(row, from_=2, to=12, width=4)
        self._web_n.set(6)
        self._web_n.pack(side=tk.LEFT, padx=4)
        self._web_btn = Chip(row, L.tr("检索并下载"), self._insp_fetch_web,
                             color=BEAD, fg="#FFFFFF", font=(FONT, 9, "bold"))
        self._web_btn.pack(side=tk.LEFT, padx=10)
        self._web_status = tk.Label(parent, text="", bg=BG, fg=TEAL,
                                    font=(FONT, 9))
        self._web_status.pack(anchor=tk.W, padx=14, pady=(2, 6))

        info = ("尊重版权与平台规则：本功能仅限个人创作参考，小批量限速下载；\n"
                "小红书 / 微信公众号需登录且有风控，未内置直连抓取。需要更多图片时：")
        tk.Label(parent, text=info, bg=BG, fg=INK_FAINT, font=(FONT, 8),
                 justify=tk.LEFT).pack(anchor=tk.W, padx=14, pady=(6, 2))
        brow = tk.Frame(parent, bg=BG)
        brow.pack(anchor=tk.W, padx=14, pady=2)
        Chip(brow, L.tr("在浏览器打开小红书搜索"), self._open_web("xiaohongshu"),
             color="#E9EDF3", fg=INK_SOFT, font=(FONT, 8, "bold")).pack(side=tk.LEFT)
        Chip(brow, L.tr("在浏览器打开百度图片"), self._open_web("baidu"),
             color="#E9EDF3", fg=INK_SOFT, font=(FONT, 8, "bold")).pack(
                 side=tk.LEFT, padx=8)
        tk.Label(parent, text=L.tr("浏览器里看到喜欢的图，可以右键保存，再回「精选·本地灵感」页点「从本地导入图片」加入灵感库。"),
                 bg=BG, fg=INK_FAINT, font=(FONT, 8), justify=tk.LEFT,
                 wraplength=620).pack(anchor=tk.W, padx=14, pady=(8, 0))

    def _open_web(self, site):
        import urllib.parse as up

        def go():
            kw = up.quote(self._web_kw.get().strip() or "拼豆图纸")
            if site == "xiaohongshu":
                url = "https://www.xiaohongshu.com/search_result?keyword=%s" % kw
            else:
                url = ("https://image.baidu.com/search/index?"
                       "tn=baiduimage&word=%s" % kw)
            os.startfile(url)
        return go

    def _insp_fetch_web(self):
        kw = self._web_kw.get().strip() or L.tr("拼豆图纸")
        try:
            n = int(self._web_n.get())
        except Exception:
            n = 6
        n = max(2, min(12, n))
        self._web_btn.set_text(L.tr("采集中…"))
        threading.Thread(target=self._insp_fetch_worker, args=(kw, n),
                         daemon=True).start()

    def _insp_fetch_worker(self, kw, n):
        try:
            res = ins.fetch_from_web([kw], limit=n)
        except Exception as e:
            res = None
            err = str(e)
        else:
            err = None
        self.after(0, lambda: self._insp_fetch_done(res, err))

    def _insp_fetch_done(self, res, err):
        self._web_btn.set_text(L.tr("检索并下载"))
        if err:
            self._web_status.configure(text="采集失败：%s" % err, fg=BEAD)
            return
        if not res:
            self._web_status.configure(text=L.tr("没有下载到可用的图片，换个关键词试试。"), fg=BEAD)
            return
        self._web_status.configure(
            text="已下载 %d 张灵感图（来源已记录），去「精选·本地灵感」页选用。" % len(res),
            fg=TEAL)
        self._insp_refresh()
        try:
            self._insp_nb.select(0)
        except Exception:
            pass

    # ---------------------------------------------------------- 记住上次设置
    def _remember_settings(self):
        """把当前设置写盘：色卡 / 颜色数 / 网格尺寸 / 勾选项 / 最近图片。"""
        try:
            _save_settings({
                "palette": self.pal_var.get(),
                "max_colors": int(self.maxC.get()) if str(self.maxC.get()).strip() else 12,
                "grid_w": int(self.lW.get()) if str(self.lW.get()).strip() else 29,
                "grid_h": int(self.lH.get()) if str(self.lH.get()).strip() else 29,
                "keep_ratio": bool(self.keep_ratio.get()),
                "bg_fill": bool(self.bg_fill.get()),
                "mode": self.mode.get(),
                "recent_image": self._last_path or "",
            "ai_cartoon_key": getattr(self, "_ai_key_var", tk.StringVar())
                                  .get() if hasattr(self, "_ai_key_var") else "",
            })
        except Exception:
            pass

    def _restore_settings(self):
        s = _load_settings()
        try:
            if s.get("grid_w"):
                self.lW.set(s["grid_w"])
            if s.get("grid_h"):
                self.lH.set(s["grid_h"])
            if s.get("max_colors"):
                self.maxC.set(s["max_colors"])
            if s.get("palette"):
                pal = self._find_palette(s["palette"])
                if pal:
                    self._refresh_palette_box(pal["name"])
                else:
                    self._refresh_palette_box()
            self.keep_ratio.set(bool(s.get("keep_ratio", False)))
            self.bg_fill.set(bool(s.get("bg_fill", False)))
            if s.get("mode") in ("auto", "manual", "none"):
                self.mode.set(s["mode"])
            recent = s.get("recent_image", "")
            if recent and os.path.exists(recent):
                self.open_image(recent, suggest=False)
            if s.get("ai_cartoon_key"):
                self._ai_key_var.set(s["ai_cartoon_key"])
        except Exception:
            pass

    def on_close(self):
        try:
            self._remember_settings()
        finally:
            self.destroy()

    # ---------------------------------------------------------- 色卡编辑
    def _find_palette(self, name):
        for p in palettes.get_all_palettes():
            if p["name"] == name:
                return p
        return None

    def _refresh_palette_box(self, keep=None):
        names = [p["name"] for p in palettes.get_all_palettes()]
        self.pal_box.configure(values=names)
        if keep is None:
            keep = self.pal_var.get()
        if keep not in names:
            keep = names[0] if names else ""
        self.pal_var.set(keep)

    def edit_palette(self):
        """色卡编辑器：自定义颜色 + 我的色卡（新建/编辑/删除/导入），持久化到 user_palettes.json。"""
        dlg = tk.Toplevel(self)
        dlg.configure(bg=BG)
        dlg.title(L.tr("色卡管理"))
        dlg.transient(self)
        dlg.resizable(False, False)

        # ---- 新增 / 删除自定义颜色 ----
        f1 = tk.LabelFrame(dlg, text=L.tr("自定义颜色（品牌色号）"), bg=BG, fg=INK,
                           font=(FONT, 10, "bold"), padx=10, pady=8)
        f1.pack(fill=tk.X, padx=14, pady=(12, 6))
        self._custom_list = tk.Listbox(f1, width=34, height=7,
                                       font=(FONT, 9), bg=PAPER, fg=INK,
                                       selectbackground="#D9E7E4")
        self._custom_list.pack(side=tk.LEFT, fill=tk.Y)
        sf = tk.Frame(f1, bg=BG)
        sf.pack(side=tk.LEFT, fill=tk.X, padx=(10, 0))
        tk.Label(sf, text=L.tr("颜色名"), bg=BG, fg=INK, font=(FONT, 9)).pack(anchor=tk.W)
        e_name = tk.Entry(sf, width=16, font=(FONT, 9))
        e_name.pack(anchor=tk.W, pady=(2, 6))
        tk.Label(sf, text=L.tr("十六进制（如 #E60012）"), bg=BG, fg=INK,
                 font=(FONT, 9)).pack(anchor=tk.W)
        e_hex = tk.Entry(sf, width=16, font=(FONT, 9))
        e_hex.pack(anchor=tk.W, pady=(2, 6))

        def add_color():
            name = e_name.get().strip()
            hx = e_hex.get().strip().lstrip("#").upper()
            if not name or not hx:
                messagebox.showinfo(L.tr("色卡"), L.tr("请填写颜色名与十六进制色值。"), parent=dlg)
                return
            try:
                int(hx, 16)
            except ValueError:
                messagebox.showinfo(L.tr("色卡"), L.tr("色值不是有效十六进制。"), parent=dlg)
                return
            if len(hx) != 6:
                messagebox.showinfo(L.tr("色卡"), L.tr("色值需要 6 位十六进制。"), parent=dlg)
                return
            palettes.save_user_data(custom_hex={**palettes.custom_colors(),
                                                name: "#" + hx})
            self._refill_custom()
            e_name.delete(0, tk.END)
            e_hex.delete(0, tk.END)

        Chip(sf, L.tr("新增颜色"), add_color, color=TEAL, fg="#FFFFFF",
             font=(FONT, 8, "bold")).pack(anchor=tk.W, pady=(2, 4))
        Chip(sf, L.tr("删除选中"), self._del_custom, color="#E9EDF3", fg=INK_SOFT,
             font=(FONT, 8, "bold")).pack(anchor=tk.W)

        # ---- 我的色卡 ----
        f2 = tk.LabelFrame(dlg, text=L.tr("我的色卡"), bg=BG, fg=INK,
                           font=(FONT, 10, "bold"), padx=10, pady=8)
        f2.pack(fill=tk.X, padx=14, pady=(4, 6))
        cols = ttk.Treeview(f2, columns=("name", "cnt"), show="headings",
                            height=6)
        cols.heading("name", text=L.tr("色卡名称"))
        cols.heading("cnt", text=L.tr("颜色数"), anchor=tk.CENTER)
        cols.column("name", width=210)
        cols.column("cnt", width=60, anchor=tk.CENTER)
        cols.pack(side=tk.LEFT, fill=tk.Y)
        sf2 = tk.Frame(f2, bg=BG)
        sf2.pack(side=tk.LEFT, fill=tk.X, padx=(10, 0))
        Chip(sf2, L.tr("新建"), lambda: self._edit_palette_colors(None, dlg),
             color=TEAL, fg="#FFFFFF", font=(FONT, 8, "bold")).pack(anchor=tk.W, pady=2)
        Chip(sf2, L.tr("编辑颜色"), lambda: self._edit_sel_palette(dlg),
             color="#E9EDF3", fg=INK_SOFT, font=(FONT, 8, "bold")).pack(anchor=tk.W, pady=2)
        Chip(sf2, L.tr("重命名"), self._rename_sel_palette, color="#E9EDF3",
             fg=INK_SOFT, font=(FONT, 8, "bold")).pack(anchor=tk.W, pady=2)
        Chip(sf2, L.tr("删除"), self._delete_palette, color=BEAD, fg="#FFFFFF",
             font=(FONT, 8, "bold")).pack(anchor=tk.W, pady=2)
        Chip(sf2, L.tr("导入 JSON"), self._import_palette, color="#E9EDF3",
             fg=INK_SOFT, font=(FONT, 8, "bold")).pack(anchor=tk.W, pady=2)

        tk.Label(dlg,
                 text="导入支持两种格式：\n"
                 "· colors 列表 —— 引用已有颜色名；\n"
                 "· hexes 映射 —— 颜色名对应真实色号（如品牌豆精确色号）。\n"
                 "所有修改点「保存」后写入 user_palettes.json。",
                 bg=BG, fg=INK_FAINT, font=(FONT, 8), justify=tk.LEFT).pack(
                     anchor=tk.W, padx=18, pady=(2, 8))

        btns = tk.Frame(dlg, bg=BG)
        btns.pack(pady=(0, 12))
        _mk_button(btns, L.tr("保 存"), lambda: self._save_palette_all(dlg)).pack(side=tk.LEFT)
        Chip(btns, L.tr("关闭"), dlg.destroy, color=CARD, fg=INK_SOFT).pack(
            side=tk.LEFT, padx=(8, 0))

        self._pal_tree = cols
        self._refill_custom()
        self._refill_palettes()
        dlg.grab_set()
        dlg.wait_window()

    # ---- 编辑器内部方法 ----
    def _refill_custom(self):
        self._custom_list.delete(0, tk.END)
        for name, hx in sorted(palettes.custom_colors().items()):
            self._custom_list.insert(tk.END, "%s  %s" % (name, hx))

    def _del_custom(self):
        sel = self._custom_list.curselection()
        if not sel:
            return
        item = self._custom_list.get(sel[0])
        name = item.split("  ")[0]
        colors = palettes.custom_colors()
        colors.pop(name, None)
        palettes.save_user_data(custom_hex=colors)
        self._refill_custom()

    def _refill_palettes(self):
        tree = self._pal_tree
        for i in tree.get_children():
            tree.delete(i)
        for p in palettes.get_all_palettes()[len(palettes.BUILT_IN_PALETTES):]:
            tree.insert("", tk.END, iid=p["name"],
                        values=(p["name"], len(p["colors"])))

    def _sel_palette(self):
        sel = self._pal_tree.selection()
        if not sel:
            messagebox.showinfo(L.tr("色卡"), L.tr("请先在下表选中一个我的色卡。"), parent=self)
            return None
        name = sel[0]
        for p in palettes.get_all_palettes():
            if p["name"] == name:
                return p
        return None

    def _edit_palette_colors(self, palette_name, parent_dlg=None):
        """勾选颜色构建/编辑一张用户色卡。palette_name=None 表示新建。"""
        dlg = tk.Toplevel(parent_dlg or self)
        dlg.configure(bg=BG)
        dlg.title("编辑色卡颜色：%s" % (palette_name or L.tr("新建")))
        dlg.transient(parent_dlg or self)

        tk.Label(dlg, text=L.tr("勾选该色卡包含的颜色（搜索可快速过滤）："),
                 bg=BG, fg=INK, font=(FONT, 9)).pack(anchor=tk.W, padx=14,
                                                     pady=(12, 4))
        search = tk.Entry(dlg, width=28, font=(FONT, 9))
        search.pack(anchor=tk.W, padx=14, pady=(0, 6))
        cur = set()
        if palette_name:
            for p in palettes.get_all_palettes():
                if p["name"] == palette_name:
                    cur = set(p["colors"])
                    break

        wrap = tk.Frame(dlg, bg=PAPER, bd=1, relief="solid")
        wrap.pack(fill=tk.BOTH, expand=True, padx=14, pady=(0, 6))
        cv = tk.Canvas(wrap, bg=PAPER, highlightthickness=0, width=380, height=300)
        sb = tk.Scrollbar(wrap, orient=tk.VERTICAL, command=cv.yview)
        body = tk.Frame(cv, bg=PAPER)
        cv.create_window((0, 0), window=body, anchor=tk.NW)
        cv.configure(yscrollcommand=sb.set)
        cv.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        body.bind("<Configure>", lambda e: cv.configure(scrollregion=cv.bbox("all")))
        cv.bind("<MouseWheel>",
                lambda e: cv.yview_scroll(-1 * (e.delta // 120), "units"))

        all_names = palettes.all_color_names()
        vars_ = {}

        def rebuild(keyword=""):
            for w in body.winfo_children():
                w.destroy()
            vars_.clear()          # 避免搜索过滤后残留旧勾选
            keyword = keyword.strip()
            cnt = 0
            for name in all_names:
                if keyword and keyword not in name:
                    continue
                var = tk.BooleanVar(value=(name in cur))
                vars_[name] = var
                chk = tk.Checkbutton(body, text=name, variable=var, bg=PAPER,
                                     fg=INK, font=(FONT, 9), anchor=tk.W,
                                     activebackground=PAPER)
                chk.grid(row=cnt // 2, column=cnt % 2, sticky=tk.W, padx=8,
                         pady=1)
                cnt += 1

        def on_search(_ev=None):
            rebuild(search.get())

        search.bind("<KeyRelease>", on_search)
        rebuild()

        btns = tk.Frame(dlg, bg=BG)
        btns.pack(pady=(0, 12))

        def apply_():
            selcol = [n for n, v in vars_.items() if v.get()]
            if not selcol:
                messagebox.showinfo(L.tr("色卡"), L.tr("至少勾选 1 个颜色。"), parent=dlg)
                return
            user = [p for p in palettes.get_all_palettes()
                    if p not in palettes.BUILT_IN_PALETTES]
            if palette_name is None:      # 新建
                name = self._ask_name(dlg)
                if not name:
                    return
                if name in [p["name"] for p in user]:
                    messagebox.showinfo(L.tr("色卡"), "已存在同名色卡：%s" % name, parent=dlg)
                    return
                user.append({"name": name, "colors": selcol})
            else:                          # 编辑已有
                for p in user:
                    if p["name"] == palette_name:
                        p["colors"] = selcol
            palettes.save_user_data(user_palettes=user)
            self.refresh_after_palette_edit()
            dlg.destroy()

        _mk_button(btns, L.tr("保 存"), apply_).pack(side=tk.LEFT)
        Chip(btns, L.tr("取消"), dlg.destroy, color=CARD, fg=INK_SOFT).pack(
            side=tk.LEFT, padx=(8, 0))
        base = parent_dlg or self
        dlg.geometry("+%d+%d" % (base.winfo_rootx() + 120,
                                 base.winfo_rooty() + 80))

    def _ask_name(self, parent):
        dlg = tk.Toplevel(parent)
        dlg.configure(bg=BG)
        dlg.title(L.tr("新建色卡"))
        dlg.transient(parent)
        dlg.resizable(False, False)
        tk.Label(dlg, text=L.tr("色卡名称："), bg=BG, fg=INK, font=(FONT, 10)).pack(
            anchor=tk.W, padx=14, pady=(12, 4))
        ent = tk.Entry(dlg, width=24, font=(FONT, 10))
        ent.pack(padx=14)
        res = {}
        btns = tk.Frame(dlg, bg=BG)
        btns.pack(pady=(10, 12))
        _mk_button(btns, L.tr("确 定"), lambda: (res.update(name=ent.get().strip()),
                                           dlg.destroy())).pack(side=tk.LEFT)
        Chip(btns, L.tr("取消"), lambda: (res.update(name=""), dlg.destroy()),
             color=CARD, fg=INK_SOFT).pack(side=tk.LEFT, padx=(8, 0))
        dlg.grab_set()
        dlg.wait_window()
        return res.get("name", "")

    def refresh_after_palette_edit(self):
        self._refill_custom()
        self._refill_palettes()
        self._refresh_palette_box()

    def _edit_sel_palette(self, parent_dlg=None):
        p = self._sel_palette()
        if p:
            self._edit_palette_colors(p["name"], parent_dlg)

    def _rename_sel_palette(self):
        p = self._sel_palette()
        if not p:
            return
        newname = self._ask_name(self)
        if not newname or newname == p["name"]:
            return
        user = [q for q in palettes.get_all_palettes()
                if q not in palettes.BUILT_IN_PALETTES]
        if newname in [q["name"] for q in user]:
            messagebox.showinfo(L.tr("色卡"), L.tr("已存在同名色卡。"))
            return
        for q in user:
            if q["name"] == p["name"]:
                q["name"] = newname
        palettes.save_user_data(user_palettes=user)
        self.refresh_after_palette_edit()

    def _delete_palette(self):
        p = self._sel_palette()
        if not p:
            return
        if not messagebox.askyesno(L.tr("色卡"), "确定删除色卡「%s」？" % p["name"]):
            return
        user = [q for q in palettes.get_all_palettes()
                if q not in palettes.BUILT_IN_PALETTES]
        user = [q for q in user if q["name"] != p["name"]]
        palettes.save_user_data(user_palettes=user)
        self.refresh_after_palette_edit()

    def _import_palette(self):
        path = filedialog.askopenfilename(title=L.tr("导入色卡 JSON"),
                                          filetypes=[("JSON", "*.json"),
                                                     (L.tr("所有文件"), "*.*")])
        if not path:
            return
        import json
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            messagebox.showerror(L.tr("色卡"), f"读取 JSON 失败：{e}")
            return
        try:
            added = palettes.import_palettes_data(data)
            if added == 0:
                messagebox.showinfo(L.tr("色卡"), L.tr("导入文件里没有可用的色卡内容。"))
                return
            self.refresh_after_palette_edit()
            messagebox.showinfo(L.tr("色卡"), "成功导入 %d 张色卡，并已持久化。" % added)
        except Exception as e:
            messagebox.showerror(L.tr("色卡"), f"导入失败：{e}")

    def _save_palette_all(self, dlg):
        palettes.save_user_data()   # 以内存当前值为准
        self.refresh_after_palette_edit()
        self.status(L.tr("色卡已保存并持久化。"))
        dlg.destroy()


def main():
    try:
        app = App()
        app.mainloop()
    except Exception:
        import traceback
        tb = traceback.format_exc()
        try:
            with open(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                   L.tr("pixel-beads_错误日志.txt")), "w", encoding="utf-8") as f:
                f.write(tb)
        except Exception:
            pass
        try:
            messagebox.showerror(APP_NAME, f"{L.tr('程序发生错误：\\n')}{tb}{L.tr('\\n\\n已写入《pixel-beads_错误日志.txt》。')}")
        except Exception:
            pass


if __name__ == "__main__":
    main()